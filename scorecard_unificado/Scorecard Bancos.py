import os
import pandas as pd
import requests
import time
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# =========================
# PARÂMETROS
# =========================
tipo_inst = 1
relatorios = {
    '2': 'Ativo',
    '3': 'Passivo',
    '4': 'DRE',
    '7': 'CarteiraTotal',
    '5': 'EstruturaDeCapital'
}

# bancos (ordem preservada)
bancos = {
    'C0083694': 'AGIBANK',
    'C0081256': 'ANDBANK',
    'C0080312': 'Banco ABC',
    'C0084480': 'BANCO BRASILEIRO',
    'C0080336': 'BTG Pactual',
    'C0080484': 'Banco BV',
    'C0081407': 'CNH Capital',
    'C0081744': 'Daycoval',
    'C0081799': 'Banco Digimais',
    'C0080329': 'BANCO DO BRASIL',
    'C0081593': 'Banco do Nordeste',
    'C0081483': 'BANCO FIBRA',
    'C0084741': 'BANCO HSBC',
    'C0080996': 'BANCO INTER',
    'C0080343': 'John Deere',
    'C0081706': 'BANCO LUSO BRASILEIRO',
    'C0080123': 'Mercantil do Brasil',
    'C0080903': 'Banco Original',
    'C0080374': 'BANCO PINE',
    'C0080109': 'BANCO SAFRA',
    'C0087298': 'BANCO STELLANTIS',
    'C0081562': 'BANCO TOYOTA DO BRASIL',
    'C0080202': 'BANCO VOLKSWAGEN',
    'C0080745': 'SICREDI',
    'C0081555': 'RABOBANK',
    'C0080178': 'Banco BMG',
    'C0084624': 'BOCOM BBM',
    'C0080075': 'Bradesco',
    'C0084844': 'Banco C6',
    'C0080738': 'Caixa',
    'C0080099': 'Itaú',
    'C0080116': 'JP MORGAN CHASE',
    'C0080460': 'OMNI',
    'C0084813': 'PAGBANK',
    'C0080855': 'BANCO RODOBENS',
    'C0080185': 'SANTANDER',
    'C0082475': 'XP Investimentos',
    'C0080848': 'Banco GM',
    'C0081328': 'BDMG',
    'C0088022': 'Picpay',
    'C0083704': 'Facta Financeira',
}

# períodos por relatório 
periodos_por_rel = {
   '4': ['202506', '202512'],   # DRE
    '2': ['202412', '202512'],   # Ativo
    '3': ['202412', '202512'],   # Passivo
    '7': ['202512'],             # Carteira de Crédito
    '5': ['202512'],             # Estrutura de Capital
}

# planilha de destino
_ONEDRIVE_RATING = os.path.join(os.path.expanduser("~"), "OneDrive - Douro Capital Gestora de Recursos e Investimentos Ltda", "Douro - Investimentos (1)", "Análise de Crédito", "Rating Crédito")
ARQUIVO_XLSM = os.path.join(_ONEDRIVE_RATING, "Watch List Bancos.xlsm")
ABA_DESTINO = "Fund Quant"
LINHA_INICIAL = 4  # começar na linha 4

# =========================
# HELPERS
# =========================
def normaliza_nome_coluna(s: str) -> str:
    """Normaliza NomeColuna para facilitar match."""
    if s is None:
        return ""
    s2 = s.replace("\r", "").replace("\n", " / ")
    s2 = " ".join(s2.split())
    return s2.strip()

def get_valor(contas_por_banco, cod_inst, chaves):
    ser = contas_por_banco.get(cod_inst, pd.Series(dtype=float))
    if ser is None or ser.empty:
        return None

    mapa = {normaliza_nome_coluna(k): v for k, v in ser.items()}

    for chave in chaves:
        alvo = normaliza_nome_coluna(chave)
        if alvo in mapa:
            return mapa[alvo]
        for k_norm, v in mapa.items():
            if k_norm.startswith(alvo):
                return v
    return None

def baixa_relatorio(relatorio_id: str, nome_rel: str, lista_periodos, session: requests.Session):
    contas_por_banco = {}
    temp_data = {cod: [] for cod in bancos.keys()}
    
    for anomes in lista_periodos:
        url = (
            "https://olinda.bcb.gov.br/olinda/servico/IFDATA/versao/v1/odata/"
            f"IfDataValores(AnoMes=@AnoMes,TipoInstituicao=@TipoInstituicao,Relatorio=@Relatorio)?"
            f"@AnoMes={anomes}&@TipoInstituicao={tipo_inst}&@Relatorio='{relatorio_id}'"
            "&$top=100000&$format=json"
            "&$select=CodInst,NomeColuna,Saldo"
        )
        
        sucesso = False
        for tentativa in range(3):
            try:
                r = session.get(url, timeout=120)
                if r.status_code == 200:
                    sucesso = True
                    break
                else:
                    print(f"❌ Erro API {nome_rel} - {anomes} (HTTP {r.status_code}) - Tentativa {tentativa+1}/3")
                    time.sleep(3)
            except requests.exceptions.RequestException as e:
                print(f"⚠️ Timeout/Erro API {nome_rel} - {anomes} ({type(e).__name__}) - Tentativa {tentativa+1}/3")
                time.sleep(5)
        
        if not sucesso:
            print(f"❌ Desistindo de {nome_rel} - {anomes} após 3 tentativas.")
            continue
        
        dados = r.json().get("value", [])
        if not dados:
            print(f"⚠️ Sem dados {nome_rel} - {anomes}")
            continue
        
        df = pd.DataFrame(dados)
        df = df[df["CodInst"].isin(bancos.keys())].copy()
        if df.empty:
            continue
        
        grp = df.groupby(["CodInst", "NomeColuna"], as_index=False)["Saldo"].sum()
        
        for cod in bancos.keys():
            ser_b = grp[grp["CodInst"] == cod].set_index("NomeColuna")["Saldo"]
            temp_data[cod].append(ser_b)
    
    for cod in bancos.keys():
        series_list = temp_data[cod]
        if not series_list:
            continue
            
        if relatorio_id in ('2', '3'):
            df_temp = pd.concat(series_list, axis=1).fillna(0)
            contas_por_banco[cod] = df_temp.mean(axis=1)
        elif relatorio_id == '4':
            df_temp = pd.concat(series_list, axis=1).fillna(0)
            contas_por_banco[cod] = df_temp.sum(axis=1)
        else:
            contas_por_banco[cod] = series_list[-1]
    
    return contas_por_banco

# =========================
# CONSTANTES DAS CONTAS
# =========================
CH_BASILEIA = ["Índice de Basileia / (n) = (e) / (j)", "Índice de Basileia (n) = (e) / (j)", "Índice de Basileia"]
CH_PL_ATIVO = ["Patrimônio Líquido / (i)", "Patrimônio Líquido (i)", "Patrimônio Líquido"]
CH_ALAVANC = ["Razão de Alavancagem / (o) = (c) / (k)", "Razão de Alavancagem (o) = (c) / (k)", "Razão de Alavancagem"]
CH_IMOB = ["Índice de Imobilização / (p)", "Índice de Imobilização (p)", "Índice de Imobilização"]
CH_PROV = ["Resultado de Provisão para Créditos de Difícil Liquidação / (b5)", "Resultado de Provisão para Créditos de Difícil Liquidação (b5)", "Resultado de Provisão para Créditos de Difícil Liquidação"]
CH_OCRED = ["Operações de Crédito / (d1)", "Operações de Crédito (d1)", "Operações de Crédito"]
CH_REC_INT = ["Receitas de Intermediação Financeira / (a)", "Receitas de Intermediação Financeira (a)", "Receitas de Intermediação Financeira"]
CH_LUCRO = ["Lucro Líquido / (j) = (g) + (h) + (i)", "Lucro Líquido (j) = (g) + (h) + (i)", "Lucro Líquido / (j)", "Lucro Líquido (j)", "Lucro Líquido"]

# Linhas ESPECÍFICAS para o Índice de Eficiência (DRE)
CH_EXP_O = ["Despesas de Pessoal / (o)", "Despesas de Pessoal (o)", "Despesas de Pessoal"]
CH_EXP_P = ["Despesas Administrativas / (p)", "Despesas Administrativas (p)", "Despesas Administrativas"]
CH_REV_IF = ["Receitas de Intermediação Financeira / (a) = (a1) + (a2) + (a3) + (a4) + (a5) + (a6)", "Receitas de Intermediação Financeira (a) = (a1) + (a2) + (a3) + (a4) + (a5) + (a6)", "Receitas de Intermediação Financeira"]
CH_RES_IF = ["Resultado de Intermediação Financeira / (k) = (a) + (b) + (c) + (d) + (e) + (f) + (g) + (h) + (i) + (j)", "Resultado de Intermediação Financeira / (c) = (a) + (b)"]
CH_REV_RPS = ["Rendas de Prestação de Serviços / (d1)", "Rendas de Prestação de Serviços (d1)", "Rendas de Prestação de Serviços"]

# =========================
# DEBUG DA EFICIÊNCIA
# =========================
def debug_eficiencia(cod, nome, contas_dre):
    print(f"\n{'='*60}")
    print(f"🏦 {nome} ({cod})")
    
    ser = contas_dre.get(cod, pd.Series(dtype=float))
    if ser is None or ser.empty:
        print("❌ Sem DRE")
        return

    # mostra TODAS as linhas relevantes
    print("\n📊 LINHAS DA DRE (filtradas):")
    for k, v in ser.items():
        k_norm = normaliza_nome_coluna(k)
        
        if any(x in k_norm.lower() for x in [
            "receita", "despesa", "intermediação", 
            "serviço", "administrativa", "pessoal"
        ]):
            print(f"{k_norm:80} | {v:,.2f}")

    # pega valores usados
    desp_pessoal = get_valor(contas_dre, cod, CH_EXP_O)
    desp_admin   = get_valor(contas_dre, cod, CH_EXP_P)

    res_if  = get_valor(contas_dre, cod, CH_RES_IF)
    rec_srv = get_valor(contas_dre, cod, CH_REV_RPS)

    print("\n🧮 COMPONENTES USADOS:")
    print(f"Despesas Pessoal:        {desp_pessoal}")
    print(f"Despesas Administrativas:{desp_admin}")
    print(f"Receita Intermediação:   {res_if}")
    print(f"Receita Serviços:        {rec_srv}")

    total_desp = None
    if desp_pessoal is not None or desp_admin is not None:
        total_desp = (desp_pessoal or 0) + (desp_admin or 0)

    total_rec = None
    if res_if is not None or rec_srv is not None:
        total_rec = (res_if or 0) + (rec_srv or 0)

    print("\n📈 RESULTADO:")
    print(f"Total Despesas: {total_desp}")
    print(f"Total Receitas: {total_rec}")

    if total_rec not in (None, 0):
        print(f"Eficiência: {abs(total_desp)/total_rec:.2%}")
    else:
        print("❌ Receita inválida")

def main():
    # =========================
    # BAIXA DADOS DOS RELATÓRIOS
    # =========================
    with requests.Session() as session:
        contas_estrut = baixa_relatorio('5', 'EstruturaDeCapital', periodos_por_rel['5'], session)
        contas_ativo  = baixa_relatorio('2', 'Ativo',               periodos_por_rel['2'], session)
        contas_pass   = baixa_relatorio('3', 'Passivo',             periodos_por_rel['3'], session)
        contas_dre    = baixa_relatorio('4', 'DRE',                 periodos_por_rel['4'], session)

    # =========================
    # MONTAGEM DOS REGISTROS
    # =========================
    registros = []

    for cod, nome in bancos.items():

        basileia = get_valor(contas_estrut, cod, CH_BASILEIA)

        pl = get_valor(contas_ativo, cod, CH_PL_ATIVO)
        if pl is None:
            pl = get_valor(contas_pass, cod, CH_PL_ATIVO)

        alav = get_valor(contas_estrut, cod, CH_ALAVANC)
        imob = get_valor(contas_estrut, cod, CH_IMOB)

        prov = get_valor(contas_dre, cod, CH_PROV)
        ocred = get_valor(contas_ativo, cod, CH_OCRED)
        prov_ratio = None
        if prov is not None and ocred not in (None, 0):
            prov_ratio = - prov / ocred

        rec_int = get_valor(contas_dre, cod, CH_REC_INT)
        lucro   = get_valor(contas_dre, cod, CH_LUCRO)

        # ===== ÍNDICE DE EFICIÊNCIA =====
        desp_pessoal = get_valor(contas_dre, cod, CH_EXP_O)
        desp_admin   = get_valor(contas_dre, cod, CH_EXP_P)
        res_if       = get_valor(contas_dre, cod, CH_RES_IF)
        rec_srv      = get_valor(contas_dre, cod, CH_REV_RPS)

        total_despesas = None
        if desp_pessoal is not None or desp_admin is not None:
            total_despesas = (desp_pessoal or 0) + (desp_admin or 0)

        total_receitas = None
        if res_if is not None or rec_srv is not None:
            total_receitas = (res_if or 0) + (rec_srv or 0)

        eficiencia = None
        if total_despesas is not None and total_receitas not in (None, 0) and total_receitas > 0:
            eficiencia = abs(total_despesas) / float(total_receitas)

        # Métricas adicionais
        lucro_sobre_receita = None
        if lucro is not None and rec_int not in (None, 0):
            lucro_sobre_receita = lucro / rec_int

        pl_passivo = get_valor(contas_pass, cod, CH_PL_ATIVO)
        if pl_passivo is None:
            pl_passivo = pl

        lucro_sobre_pl = None
        if lucro is not None and pl_passivo not in (None, 0):
            lucro_sobre_pl = lucro / pl_passivo

        registros.append([
            nome,                      # A
            cod,                       # B
            basileia,                  # C
            pl,                        # D (R$)
            alav,                      # E
            imob,                      # F
            prov_ratio,                # G
            eficiencia,                # H
            lucro_sobre_receita,       # I
            lucro_sobre_pl             # J
        ])

    # =========================
    # ESCREVE NO EXCEL (XLSM)
    # =========================
    wb = load_workbook(ARQUIVO_XLSM, keep_vba=True)
    if ABA_DESTINO not in wb.sheetnames:
        raise ValueError(f"A aba '{ABA_DESTINO}' não existe no arquivo alvo.")

    ws = wb[ABA_DESTINO]

    # Cabeçalhos
    fundquant_cols = [
        "Nome do Banco",
        "Código do Banco",
        "Indice de Basileia",
        "PL",
        "Indice de alavancagem",
        "Indice de imobilização",
        "Despesa de provisão sobre carteira",
        "Índice de Eficiência",
        "Margem Liquida",
        "ROE"
    ]

    for col_num, header in enumerate(fundquant_cols, 1):
        ws.cell(row=LINHA_INICIAL-1, column=col_num, value=header)

    # Linhas
    linha = LINHA_INICIAL
    for (nome, cod, c, d, e, f, g, h, i, j) in registros:
        ws.cell(row=linha, column=1,  value=nome)
        ws.cell(row=linha, column=2,  value=cod)
        ws.cell(row=linha, column=3,  value=c)
        ws.cell(row=linha, column=4,  value=d)
        ws.cell(row=linha, column=5,  value=e)
        ws.cell(row=linha, column=6,  value=f)
        ws.cell(row=linha, column=7,  value=g)
        ws.cell(row=linha, column=8,  value=h)
        ws.cell(row=linha, column=9,  value=i)
        ws.cell(row=linha, column=10, value=j)
        linha += 1

    # Formatação
    percent_cols = ['C', 'E', 'F', 'G', 'H', 'I', 'J']
    for col in percent_cols:
        for row in range(LINHA_INICIAL, linha):
            ws[f"{col}{row}"].number_format = '0.00%'

    for row in range(LINHA_INICIAL, linha):
        ws[f"D{row}"].number_format = 'R$ #,##0;R$ -#,##0'

    wb.save(ARQUIVO_XLSM)
    wb.close()
    print("\n✅ Dados escritos com sucesso na aba especificada.")

if __name__ == "__main__":
    main()