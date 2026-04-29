import openpyxl, os

path = r"C:\Users\Brno Goes\OneDrive - Douro Capital Gestora de Recursos e Investimentos Ltda\Douro - Investimentos\Análise de Crédito\Rating Crédito\Scorecard de Empresas.xlsm"

wb = openpyxl.load_workbook(path, data_only=True)
ws = wb["Fund Quant"]

print("=== Primeiras 6 linhas do Fund Quant (OneDrive) ===")
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i >= 6:
        break
    print(f"Linha {i+1}: {[str(v)[:20] if v is not None else None for v in row[:18]]}")

print()
print("=== Aba Ranking — primeiras linhas ===")
ws_r = wb["Ranking"]
for i, row in enumerate(ws_r.iter_rows(values_only=True)):
    if i >= 8:
        break
    print(f"Linha {i+1}: {[str(v)[:20] if v is not None else None for v in row[:14]]}")
