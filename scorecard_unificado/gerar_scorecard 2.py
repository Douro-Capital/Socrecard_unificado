"""
===============================================================
  DOURO CAPITAL — Gerador Automático do Scorecard de Crédito
  Lê planilhas .xlsm → extrai dados → gera scorecard.html estático

  Verticais:
    1. Crédito Privado  (Empresas Corporativas)
    2. Crédito Bancário (Instituições Financeiras)

  Dependências:
      pip install openpyxl jinja2

  Uso:
      python gerar_scorecard.py

  O arquivo scorecard.html será salvo na mesma pasta do script.
===============================================================
"""

import os
import json
import sys
import datetime
import base64
import openpyxl
from jinja2 import Template

# ─────────────────────────────────────────────────────────────
# CONFIGURAÇÃO — caminhos resolvidos dinamicamente com base 
# na localização relativa deste script.
# ─────────────────────────────────────────────────────────────

_SCORECARD_DIR = os.path.dirname(os.path.abspath(__file__))
_ONEDRIVE_RATING = os.path.join(os.path.expanduser("~"), "OneDrive - Douro Capital Gestora de Recursos e Investimentos Ltda", "Douro - Investimentos", "Análise de Crédito", "Rating Crédito")

EXCEL_PATH_PRIVADO = os.path.join(_ONEDRIVE_RATING, "Scorecard de Empresas.xlsm")
EXCEL_PATH_BANCOS  = os.path.join(_ONEDRIVE_RATING, "Watch List Bancos.xlsm")
OUTPUT_HTML        = os.path.join(_ONEDRIVE_RATING, "scorecard.html")
LOGO_PATH          = os.path.join(_SCORECARD_DIR, "Douro-Capital-logo-Horizontal-colorida (3).png")
# ─────────────────────────────────────────────────────────────


# ─────────────────────────────────────────────────────────────
# HELPERS COMUNS
# ─────────────────────────────────────────────────────────────

def fmt_num(v, decimals=2):
    """Converte para float arredondado, ou None se inválido."""
    if v is None:
        return None
    try:
        return round(float(v), decimals)
    except (TypeError, ValueError):
        return None


def image_to_data_uri(image_path: str) -> str:
    """Converte imagem local para data URI base64, para HTML portátil."""
    if not image_path or not os.path.exists(image_path):
        return ""
    try:
        with open(image_path, "rb") as f:
            encoded = base64.b64encode(f.read()).decode("ascii")
        ext = os.path.splitext(image_path)[1].lower()
        mime = "image/png" if ext == ".png" else "image/jpeg" if ext in (".jpg", ".jpeg") else "application/octet-stream"
        return f"data:{mime};base64,{encoded}"
    except Exception:
        return ""


# ─────────────────────────────────────────────────────────────
# VERTICAL 1: CRÉDITO PRIVADO (código original intacto)
# ─────────────────────────────────────────────────────────────

def extract_ranking(ws) -> dict:
    """Extrai aba Ranking → dict {empresa: {...}}"""
    data = {}
    for row in list(ws.iter_rows(values_only=True))[5:]:
        if row[3] is None:
            continue
        empresa = str(row[3]).strip()
        data[empresa] = {
            "rankingAprovados": row[1],
            "ranking":          row[2],
            "empresa":          empresa,
            "fiador":           str(row[4]).strip() if row[4] else None,
            "codigo":           str(row[5]).strip() if row[5] else "",
            "periodicidade":    row[6],
            "status":           row[7],
            "scoreQuantitativo": round(float(row[8]) * 100, 1) if row[8] else 0,
            "scoreQualitativo":  round(float(row[9]) * 100, 1) if row[9] else 0,
            "scoreTotal":        round(float(row[10]) * 100, 1) if row[10] else 0,
            "rating":            row[11],
        }
    return data


def extract_quant(ws) -> dict:
    """Extrai aba Fund Quant → dict {empresa: {indicadores}}.

    Suporta dois layouts do Excel:
      Layout A (Downloads): col0=Nome, col1=Código, col2=Nomencl, col3=Fiador, indicadores em col4+
      Layout B (OneDrive):  col0=Código, col1=Nomencl, col2=Fiador, indicadores em col3+
    A detecção é feita pelo cabeçalho: se col0 do header for 'Código', usa Layout B.
    """
    rows = list(ws.iter_rows(values_only=True))

    # Localiza a linha de cabeçalho (contém 'Div Liq' ou 'EBITDA')
    header_idx = 2  # default
    for i, row in enumerate(rows[:6]):
        cells = [str(c).strip().lower() if c else "" for c in row]
        if any("div liq" in c or "ebitda" in c for c in cells):
            header_idx = i
            break

    header = [str(c).strip().lower() if c else "" for c in rows[header_idx]]

    # Layout B: col0 do cabeçalho começa com 'c' de 'código' e col3 tem indicador
    if header[0].startswith("c") and ("div" in header[3] or "ebitda" in header[3]):
        # col0=Código, col1=Nomencl, col2=Fiador(nome), indicadores em col3+
        nome_col  = 2
        ind_start = 3
    else:
        # Layout A: col0=Nome, indicadores em col4+
        nome_col  = 0
        ind_start = 4

    data = {}
    for row in rows[header_idx + 1:]:
        if row[nome_col] is None:
            continue
        empresa = str(row[nome_col]).strip()
        if not empresa or empresa.lower() in ("empresa", "fiador", "none"):
            continue
        o = ind_start  # offset
        data[empresa] = {
            "div_liq_ebitda":    fmt_num(row[o]),
            "ebitda_desp_fin":   fmt_num(row[o + 1]),
            "estrutura_capital": fmt_num(row[o + 2], 3),
            "div_cp_n":          fmt_num(row[o + 3], 0),
            "div_cp_n1":         fmt_num(row[o + 4], 0),
            "evol_div_cp":       fmt_num(row[o + 5], 3),
            "liquidez_corrente": fmt_num(row[o + 6]),
            "div_cp_div_total":  fmt_num(row[o + 7], 3),
            "margem_liquida":    fmt_num(row[o + 8], 3),
            "roic":              fmt_num(row[o + 9], 3),
            "disponibilidades":  fmt_num(row[o + 10], 0),
            "caixa_div_cp":      fmt_num(row[o + 11]),
        }
    return data


def extract_qual(ws) -> dict:
    """Extrai aba Fund Qual → dict {empresa: {scores + comentários + setor}}"""
    rows = list(ws.iter_rows(values_only=True))
    data = {}
    i = 3

    def clean(v):
        s = str(v).strip() if v else None
        return None if s in ("Comentários", "None", "", "nan") else s

    while i < len(rows):
        row = rows[i]
        if not row[0] or row[0] == "Empresa":
            i += 1
            continue
        empresa = str(row[0]).strip()
        cmts = rows[i + 1] if i + 1 < len(rows) else ([None] * 10)

        data[empresa] = {
            "vantagens_score":        row[3],
            "oportunidades_score":    row[4],
            "vulnerabilidades_score": row[5],
            "governanca_score":       row[6],
            "setor_score":            row[7] if isinstance(row[7], (int, float)) else None,
            "rating_score":           row[8] if len(row) > 8 else None,
            "cmt_vantagens":          clean(cmts[3]) if len(cmts) > 3 else None,
            "cmt_oportunidades":      clean(cmts[4]) if len(cmts) > 4 else None,
            "cmt_vulnerabilidades":   clean(cmts[5]) if len(cmts) > 5 else None,
            "cmt_governanca":         clean(cmts[6]) if len(cmts) > 6 else None,
            "setor":                  clean(cmts[7]) if len(cmts) > 7 else None,
            "rating_text":            clean(cmts[8]) if len(cmts) > 8 else None,
        }
        i += 2
    return data


def build_dataset(excel_path: str) -> list:
    """Carrega o .xlsm de Crédito Privado e retorna lista de empresas com todos os dados."""
    print(f"  Abrindo Crédito Privado: {excel_path}")
    wb = openpyxl.load_workbook(excel_path, data_only=True)

    ranking = extract_ranking(wb["Ranking"])
    quant   = extract_quant(wb["Fund Quant"])
    qual    = extract_qual(wb["Fund Qual"])

    merged = []
    for emp, d in ranking.items():
        q  = quant.get(emp) or quant.get(d.get("fiador", "")) or {}
        ql = qual.get(emp)  or qual.get(d.get("fiador", ""))  or {}
        d["quant"] = q
        d["qual"]  = ql
        merged.append(d)

    merged.sort(key=lambda x: x["ranking"] or 9999)
    print(f"  {len(merged)} empresas corporativas extraídas.")
    return merged


# ─────────────────────────────────────────────────────────────
# VERTICAL 2: CRÉDITO BANCÁRIO
# ─────────────────────────────────────────────────────────────

# Critérios quantitativos de pontuação (0–5) para bancos
# Formato: (limiar_0, limiar_1, ..., limiar_5)
# Basileia: maior = melhor → pontos 0..5 conforme thresholds crescentes
# PL: maior = melhor
# Alavancagem (índice): maior = melhor
# Imobilização: menor = melhor (> limiar = pior)
# Provisão sobre carteira: menor = melhor
# Eficiência: menor = melhor (índice < 1 = bom; > 1 = ruim para receita/despesa)
# Margem Líquida: maior = melhor
# ROE: menor = melhor? Não — maior ROE = melhor

# Direção: "asc" → maior valor = maior score; "desc" → menor valor = maior score
BANK_QUANT_CRITERIA = {
    # col_idx, direção, thresholds para scores 0,1,2,3,4,5
    "basileia":   ("asc",  [0.11, 0.12, 0.13, 0.14, 0.15, 0.16]),
    "pl":         ("asc",  [5e6,  15e6, 30e6, 50e6, 100e6, 1e9]),
    "alavancagem":("asc",  [0.03, 0.05, 0.08, 0.10, 0.15, 0.20]),
    "imobilizacao":("desc",[0.40, 0.35, 0.30, 0.25, 0.20, 0.15]),
    "provisao":   ("desc", [0.07, 0.055,0.04, 0.025,0.02, 0.01]),
    "eficiencia": ("desc", [0.60, 0.55, 0.50, 0.45, 0.40, 0.30]),
    "margem":     ("asc",  [-0.05,0.0,  0.05, 0.10, 0.15, 0.20]),
    "roe":        ("asc",  [0.0,  0.05, 0.075,0.10, 0.15, 0.20]),
}

RATING_SCORE_MAP = {
    "AAA": 5, "AA+": 5, "AA": 5, "AA-": 4,
    "A+": 4, "A": 4, "A-": 3,
    "BBB+": 3, "BBB": 3, "BBB-": 2,
    "BB+": 2, "BB": 2, "BB-": 1,
    "B+": 1, "B": 1, "B-": 0,
}


def _quant_score(value, direction, thresholds):
    """Retorna score 0–5 para um indicador bancário com base nos limiares."""
    if value is None:
        return 0
    score = 0
    if direction == "asc":
        for t in thresholds:
            if value >= t:
                score += 1
    else:  # desc: menor = melhor
        for t in thresholds:
            if value <= t:
                score += 1
    return min(score, 5)


def _rating_to_canonical(rating_txt):
    """Extrai o rating canônico de strings como 'AAA com perspectiva estável'."""
    if not rating_txt:
        return None
    for r in ["AAA", "AA+", "AA-", "AA", "A+", "A-", "A",
              "BBB+", "BBB-", "BBB", "BB+", "BB-", "BB",
              "B+", "B-", "B", "CCC", "CC", "C", "D"]:
        if rating_txt.upper().startswith(r):
            return r
    return None


def extract_bank_quant(ws) -> dict:
    """
    Extrai aba 'Fund Quant' da planilha de bancos.
    Colunas (0-indexed, base linha 4 = row index 3):
      0: Nome do Banco
      1: Código
      2: Índice de Basileia
      3: PL
      4: Índice de alavancagem
      5: Índice de imobilização
      6: Despesa de provisão sobre carteira
      7: Índice de Eficiência
      8: Margem Líquida
      9: ROE
     14: Pontuação Quant  (fórmula — pode ser None)
     15: Pontuação Quali  (fórmula — pode ser None)
     16: Pontuação Total  (fórmula — pode ser None)
     18: Status
    """
    data = {}
    for row in list(ws.iter_rows(values_only=True))[3:]:
        nome = row[0]
        if not nome or not isinstance(nome, str):
            continue
        nome = nome.strip()

        basileia    = fmt_num(row[2], 4)
        pl          = fmt_num(row[3], 0)
        alavancagem = fmt_num(row[4], 4)
        imobilizacao= fmt_num(row[5], 4)
        provisao    = fmt_num(row[6], 4)
        eficiencia  = fmt_num(row[7], 4)
        margem      = fmt_num(row[8], 4)
        roe         = fmt_num(row[9], 4)
        status      = str(row[18]).strip() if row[18] else "Em análise"
        codigo      = str(row[1]).strip() if row[1] else ""

        # Calcula score quantitativo (média dos 8 indicadores, escala 0-100)
        scores_ind = {
            "basileia":    _quant_score(basileia,     "asc",  BANK_QUANT_CRITERIA["basileia"][1]),
            "pl":          _quant_score(pl,           "asc",  BANK_QUANT_CRITERIA["pl"][1]),
            "alavancagem": _quant_score(alavancagem,  "asc",  BANK_QUANT_CRITERIA["alavancagem"][1]),
            "imobilizacao":_quant_score(imobilizacao, "desc", BANK_QUANT_CRITERIA["imobilizacao"][1]),
            "provisao":    _quant_score(provisao,     "desc", BANK_QUANT_CRITERIA["provisao"][1]),
            "eficiencia":  _quant_score(eficiencia,   "desc", BANK_QUANT_CRITERIA["eficiencia"][1]),
            "margem":      _quant_score(margem,       "asc",  BANK_QUANT_CRITERIA["margem"][1]),
            "roe":         _quant_score(roe,           "asc",  BANK_QUANT_CRITERIA["roe"][1]),
        }
        score_quant_raw = sum(scores_ind.values()) / (5 * len(scores_ind))  # 0-1

        data[nome] = {
            "codigo":      codigo,
            "status":      status,
            "basileia":    basileia,
            "pl":          pl,
            "alavancagem": alavancagem,
            "imobilizacao":imobilizacao,
            "provisao":    provisao,
            "eficiencia":  eficiencia,
            "margem":      margem,
            "roe":         roe,
            "_scores_ind": scores_ind,
            "_score_quant_raw": score_quant_raw,
        }
    return data


def extract_bank_qual(ws, sector_map: dict) -> dict:
    """
    Extrai aba 'Fund Qual' da planilha de bancos.
    Estrutura: linhas pares = scores; linhas ímpares = comentários.
    Colunas (0-indexed):
      0: Empresa (ou 'Comentários' para linhas de detalhe)
      2: Vantagens comparativas (0-5)
      3: Oportunidades (0-5)
      4: Vulnerabilidades (0-5) — atenção: menor = pior, maior = melhor aqui
      5: Ameaças (não pontuado diretamente no score)
      6: Governança (0-5)
      7: Serviços (0-5)
      8: Rating (texto ou score 0-5)
    """
    rows = list(ws.iter_rows(values_only=True))
    data = []

    def clean(v):
        s = str(v).strip() if v else None
        return None if s in ("Comentários", "None", "", "nan") else s

    def is_number(v):
        return isinstance(v, (int, float))

    def is_score_row(row):
        # Aceita os dois layouts já observados na planilha:
        # 1) col 0 vazia e scores a partir da col 2
        # 2) col 0 com nome do banco e scores a partir da col 2
        if not row or len(row) < 9:
            return False
        return is_number(row[2]) and (
            row[0] is None or (isinstance(row[0], str) and str(row[0]).strip() not in ("", "Empresa", "Comentários"))
        )

    def is_comment_row(row):
        if not row or len(row) == 0:
            return False
        first = row[0]
        return isinstance(first, str) and first.strip() == "Comentários"

    i = 3  # pula header (3 linhas)
    while i < len(rows):
        row = rows[i]
        if not is_score_row(row):
            i += 1
            continue

        next_row = rows[i + 1] if i + 1 < len(rows) else [None] * 20
        cmt_row = next_row if is_comment_row(next_row) else [None] * 20

        vantagens         = row[2] if is_number(row[2]) else 0
        oportunidades     = row[3] if is_number(row[3]) else 0
        vulnerabilidades  = row[4] if is_number(row[4]) else 0
        governanca        = row[6] if is_number(row[6]) else 0
        rating_raw        = row[8]

        rating_txt_score = None
        if is_number(rating_raw):
            rating_score_val = int(rating_raw)
        else:
            rating_score_val = RATING_SCORE_MAP.get(_rating_to_canonical(str(rating_raw)) or "", 0)
            rating_txt_score = clean(rating_raw)

        # Score qualitativo: média simples das dimensões (0-5), sem inversão.
        qual_raw = (vantagens + oportunidades + vulnerabilidades + governanca + rating_score_val) / (5 * 5)

        banco_entry = {
            "vantagens_score":        vantagens,
            "oportunidades_score":    oportunidades,
            "vulnerabilidades_score": vulnerabilidades,
            "governanca_score":       governanca,
            "rating_score":           rating_score_val,
            "rating_txt":             rating_txt_score,
            "cmt_vantagens":          clean(cmt_row[2]) if len(cmt_row) > 2 else None,
            "cmt_oportunidades":      clean(cmt_row[3]) if len(cmt_row) > 3 else None,
            "cmt_vulnerabilidades":   clean(cmt_row[4]) if len(cmt_row) > 4 else None,
            "cmt_governanca":         clean(cmt_row[6]) if len(cmt_row) > 6 else None,
            "cmt_rating":             clean(cmt_row[8]) if len(cmt_row) > 8 else None,
            "_qual_raw":              qual_raw,
        }
        data.append(banco_entry)
        i += 2 if is_comment_row(next_row) else 1

    # Remontar com nomes reais (do sector_map ou da ordem de Fund Quant)
    # Retorna lista ordenada para parear com fund_quant por ordem de inserção
    return data


def extract_sector_map(ws) -> dict:
    """Extrai mapa {nome_banco: setor} da aba 'Critérios de pontuação'."""
    rows = list(ws.iter_rows(values_only=True))
    sector_map = {}
    for row in rows[2:]:
        if len(row) > 14 and row[13] is not None and row[14] is not None:
            nome = str(row[13]).strip()
            setor = str(row[14]).strip()
            if nome and setor:
                sector_map[nome] = setor
    return sector_map


def _derive_rating(score_total_pct: float) -> str:
    """Mapeia score total (0-100) → rating Douro para bancos."""
    s = score_total_pct
    if s >= 85:   return "AAA"
    if s >= 75:   return "AA"
    if s >= 65:   return "A"
    if s >= 55:   return "BBB"
    if s >= 45:   return "BB"
    if s >= 35:   return "B"
    return "CCC"


def build_bank_dataset(excel_path: str) -> list:
    """
    Carrega o .xlsm de Crédito Bancário e retorna lista de bancos
    com todos os dados, scores calculados programaticamente.
    """
    print(f"  Abrindo Crédito Bancário: {excel_path}")
    wb = openpyxl.load_workbook(excel_path, data_only=True)

    quant_data  = extract_bank_quant(wb["Fund Quant"])
    sector_map  = extract_sector_map(wb["Critérios de pontuação"])
    qual_list   = extract_bank_qual(wb["Fund Qual"], sector_map)

    # Fund Qual não tem nomes — pareamos pela ordem de aparição com Fund Quant
    quant_keys = list(quant_data.keys())  # mesma ordem de leitura

    merged = []
    for idx, (nome, qd) in enumerate(quant_data.items()):
        ql = qual_list[idx] if idx < len(qual_list) else {}

        sq_pct  = round(qd["_score_quant_raw"] * 100, 1)
        sql_pct = round(ql.get("_qual_raw", 0) * 100, 1)
        st_pct  = round((qd["_score_quant_raw"] * 0.5 + ql.get("_qual_raw", 0) * 0.5) * 100, 1)

        rating  = _derive_rating(st_pct)

        merged.append({
            "ranking":            idx + 1,   # placeholder; reordenamos abaixo
            "empresa":            nome,
            "codigo":             qd["codigo"],
            "periodicidade":      "Trimestral",
            "status":             qd["status"],
            "scoreQuantitativo":  sq_pct,
            "scoreQualitativo":   sql_pct,
            "scoreTotal":         st_pct,
            "rating":             rating,
            "setor":              sector_map.get(nome, None),
            "quant": {
                "basileia":    qd["basileia"],
                "pl":          qd["pl"],
                "alavancagem": qd["alavancagem"],
                "imobilizacao":qd["imobilizacao"],
                "provisao":    qd["provisao"],
                "eficiencia":  qd["eficiencia"],
                "margem":      qd["margem"],
                "roe":         qd["roe"],
            },
            "qual": {
                "vantagens_score":        ql.get("vantagens_score", 0),
                "oportunidades_score":    ql.get("oportunidades_score", 0),
                "vulnerabilidades_score": ql.get("vulnerabilidades_score", 0),
                "governanca_score":       ql.get("governanca_score", 0),
                "rating_score":           ql.get("rating_score", 0),
                "rating_txt":             ql.get("rating_txt"),
                "cmt_vantagens":          ql.get("cmt_vantagens"),
                "cmt_oportunidades":      ql.get("cmt_oportunidades"),
                "cmt_vulnerabilidades":   ql.get("cmt_vulnerabilidades"),
                "cmt_governanca":         ql.get("cmt_governanca"),
                "cmt_rating":             ql.get("cmt_rating"),
                "setor":                  sector_map.get(nome, None),
            },
        })

    # Ordena por score total decrescente e reatribui ranking
    # Bancos FGC (não comparáveis diretamente) ficam depois dos Aprovados
    STATUS_ORDER = {"Aprovado": 0, "Em análise": 1, "Reprovado": 2, "FGC": 3}
    merged.sort(key=lambda x: (STATUS_ORDER.get(x["status"], 9), -x["scoreTotal"]))
    for i, b in enumerate(merged):
        b["ranking"] = i + 1

    print(f"  {len(merged)} bancos extraídos.")
    return merged


# ─────────────────────────────────────────────────────────────
# TEMPLATE JINJA2 — HTML COMPLETO
# ─────────────────────────────────────────────────────────────

HTML_TEMPLATE = r"""<!DOCTYPE html>
<html lang="pt-BR">
<head>
<meta charset="UTF-8"/>
<meta name="viewport" content="width=device-width, initial-scale=1.0"/>
<title>Scorecard de Crédito — Douro Capital</title>
<link rel="preconnect" href="https://fonts.googleapis.com"/>
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin/>
<link href="https://fonts.googleapis.com/css2?family=Montserrat:wght@300;400;500;600;700;800&family=DM+Mono:wght@300;400;500;600;700&display=swap" rel="stylesheet"/>
<script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.1/chart.umd.min.js"></script>
<style>
/* ─── GERADO AUTOMATICAMENTE EM {{ generated_at }} ─── */
*,*::before,*::after{box-sizing:border-box;margin:0;padding:0}
:root{
  --navy:#1f2839;--teal:#00677b;--gold:#b69d74;--offwhite:#d5d8c9;
  --bg:#f4f5f0;--bg2:#eceee7;--surface:#ffffff;--surface2:#f8f9f5;
  --text:#1f2839;--text2:#4a5568;--text3:#718096;
  --border:#dde0d8;--border2:#c8ccbf;
  --c-aaa:#00677b;--c-aaa-bg:#e6f2f4;--c-aaa-bdr:#b3d9df;
  --c-aa:#0d8a7a;--c-aa-bg:#e6f5f3;--c-aa-bdr:#b3dfd9;
  --c-a:#7a6518;--c-a-bg:#faf5e8;--c-a-bdr:#e8d9a8;
  --c-bbb:#9a6e00;--c-bbb-bg:#fef9ec;--c-bbb-bdr:#f0d88a;
  --c-bb:#b54a00;--c-bb-bg:#fff3ec;--c-bb-bdr:#f0bfa0;
  --c-b:#c0392b;--c-b-bg:#fdf0ee;--c-b-bdr:#f0b8b3;
  --c-low:#922b21;--c-low-bg:#fce8e6;--c-low-bdr:#e8aeaa;
  --font:'Montserrat',system-ui,sans-serif;
  --mono:'DM Mono','Courier New',monospace;
}
html{scroll-behavior:smooth}
body{background:var(--bg);color:var(--text);font-family:var(--font);overflow-x:hidden;cursor:none;min-height:100vh}
#cur{position:fixed;width:7px;height:7px;background:var(--navy);border-radius:50%;pointer-events:none;z-index:9999;transform:translate(-50%,-50%);mix-blend-mode:multiply}
#cur-ring{position:fixed;width:28px;height:28px;border:1.5px solid var(--teal);border-radius:50%;pointer-events:none;z-index:9998;transform:translate(-50%,-50%);opacity:.4;transition:width .22s,height .22s}
body::before{content:'';position:fixed;inset:0;background-image:url("data:image/svg+xml,%3Csvg viewBox='0 0 200 200' xmlns='http://www.w3.org/2000/svg'%3E%3Cfilter id='n'%3E%3CfeTurbulence type='fractalNoise' baseFrequency='0.9' numOctaves='4' stitchTiles='stitch'/%3E%3C/filter%3E%3Crect width='100%25' height='100%25' filter='url(%23n)' opacity='0.025'/%3E%3C/svg%3E");pointer-events:none;z-index:0;opacity:.3}
#canvas{position:fixed;inset:0;pointer-events:none;z-index:0;opacity:.18}
#progress{position:fixed;top:0;left:0;height:2px;background:linear-gradient(90deg,var(--teal),var(--gold));z-index:1000;width:0%}

/* ═══ GLOBAL TAB TOGGLE & SEARCH ═══ */
nav{position:sticky;top:0;z-index:200;background:rgba(244,245,240,.95);backdrop-filter:blur(16px);border-bottom:1px solid var(--border);padding:0 48px}
.nav-inner{max-width:1400px;margin:0 auto;display:flex;align-items:center;justify-content:space-between;height:60px}
.nav-brand{display:flex;align-items:center;gap:12px;flex:1}
.nav-divider{width:1px;height:14px;background:var(--border2)}
.nav-subtitle{font-size:.62rem;letter-spacing:.12em;color:var(--teal);font-weight:500;text-transform:uppercase}

.nav-tabs{display:flex;align-items:center;gap:4px;background:var(--bg2);border:1px solid var(--border);border-radius:8px;padding:3px; justify-content:center;}
.nav-tab-btn{background:transparent;border:none;font-family:var(--font);font-size:.65rem;font-weight:600;letter-spacing:.1em;text-transform:uppercase;padding:6px 16px;border-radius:6px;cursor:pointer;color:var(--text3);transition:all .22s;white-space:nowrap}
.nav-tab-btn.active{background:var(--navy);color:#fff;box-shadow:0 2px 8px rgba(31,40,57,.18)}
.nav-tab-btn:hover:not(.active){color:var(--teal)}

.nav-search-wrap { position:relative; display:flex; align-items:center; flex:1; justify-content:flex-end;}
.nav-search-box { display:flex; align-items:center; background:var(--surface); border:1px solid var(--border); border-radius:8px; padding:6px 12px; width:280px; transition:border-color .2s, box-shadow .2s; cursor:text; }
.nav-search-box:focus-within { border-color:var(--teal); box-shadow:0 0 0 3px rgba(0,103,123,.1); }
.ns-icon { width:14px; height:14px; color:var(--text3); margin-right:8px; }
.ns-input { border:none; outline:none; background:transparent; font-family:var(--font); font-size:.7rem; font-weight:500; color:var(--text); width:100%; }
.ns-input::placeholder { color:var(--text3); font-weight:400; }
.ns-hint { display:flex; align-items:center; gap:4px; margin-left:8px; }
.ns-key { border:1px solid var(--border2); border-radius:4px; font-family:var(--mono); font-size:.55rem; color:var(--text3); padding:2px 6px; background:var(--surface2); }
.ns-dropdown { position:absolute; top:calc(100% + 8px); right:0; width:340px; background:var(--surface); border:1px solid var(--border); border-radius:10px; box-shadow:0 12px 40px rgba(31,40,57,.15); z-index:900; max-height:360px; overflow-y:auto; opacity:0; pointer-events:none; transform:translateY(-10px); transition:all .2s; }
.ns-dropdown.open { opacity:1; pointer-events:auto; transform:translateY(0); }
.ns-item { display:flex; align-items:center; justify-content:space-between; padding:12px 16px; border-bottom:1px solid var(--surface2); cursor:pointer; transition:background .15s; }
.ns-item:last-child { border-bottom:none; }
.ns-item:hover { background:var(--surface2); }
.ns-item-main { display:flex; flex-direction:column; gap:4px; }
.ns-item-name { font-size:.75rem; font-weight:700; color:var(--navy); }
.ns-item-meta { display:flex; align-items:center; gap:8px; font-family:var(--mono); font-size:.6rem; color:var(--text3); }
.ns-tag { font-family:var(--font); font-size:.5rem; font-weight:700; letter-spacing:.05em; text-transform:uppercase; padding:2px 6px; border-radius:4px; border:1px solid; }
.ns-tag-corp { color:var(--teal); background:rgba(0,103,123,.1); border-color:rgba(0,103,123,.2); }
.ns-tag-banco { color:var(--gold); background:rgba(182,157,116,.1); border-color:rgba(182,157,116,.2); }
.ns-status-wrap { display:flex; align-items:center; gap:4px; }

/* ═══ HERO & SECTIONS ═══ */
#hero-privado, #hero-bancos{position:relative;z-index:1;max-width:1400px;margin:0 auto;padding:56px 48px 40px}
.hero-label{font-family:var(--mono);font-size:.6rem;letter-spacing:.3em;color:var(--teal);text-transform:uppercase;margin-bottom:12px;opacity:0;animation:fadeUp .6s .2s forwards}
.hero-title{font-size:clamp(2rem,4vw,3.2rem);font-weight:700;letter-spacing:-.01em;line-height:1.1;color:var(--navy);opacity:0;animation:fadeUp .8s .35s forwards}
.hero-rule{width:48px;height:2px;background:var(--gold);margin:20px 0;opacity:0;animation:fadeUp .5s .5s forwards}
.stats-row{display:flex;gap:16px;margin-top:32px;flex-wrap:wrap;opacity:0;animation:fadeUp .7s .7s forwards}
.stat-box{background:var(--surface);border:1px solid var(--border);border-radius:8px;padding:16px 20px;position:relative;overflow:hidden}
.stat-box::after{content:'';position:absolute;bottom:0;left:0;right:0;height:2px}
.stat-box.s-total::after{background:var(--navy)}.stat-box.s-aprov::after{background:var(--teal)}.stat-box.s-anal::after{background:var(--gold)}.stat-box.s-reprov::after{background:var(--c-low)}
.stat-num{font-size:1.8rem;font-weight:700;line-height:1;color:var(--navy);letter-spacing:-.02em}
.stat-num.c-teal{color:var(--teal)}.stat-num.c-gold{color:var(--gold)}.stat-num.c-red{color:var(--c-low)}
.stat-lbl{font-size:.6rem;font-weight:600;letter-spacing:.12em;text-transform:uppercase;color:var(--text3);margin-top:4px}
.section{position:relative;z-index:1;max-width:1400px;margin:0 auto;padding:0 48px 56px}
.sec-eyebrow{font-family:var(--mono);font-size:.58rem;letter-spacing:.25em;color:var(--teal);text-transform:uppercase;margin-bottom:8px}
.sec-title{font-size:1.1rem;font-weight:700;letter-spacing:-.01em;color:var(--navy);margin-bottom:24px}
.sec-title span{color:var(--gold)}
.top5-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(220px,1fr));gap:14px}
.pod{background:var(--surface);border:1px solid var(--border);border-radius:10px;padding:20px;cursor:pointer;transition:transform .2s,box-shadow .2s,border-color .2s;position:relative;overflow:hidden}
.pod:hover,.pod.rank-1:hover,.pod.rank-2:hover,.pod.rank-3:hover{transform:translateY(-3px);box-shadow:0 12px 32px rgba(31,40,57,.12);border-color:var(--teal)}
.pod.rank-1{border-color:rgba(182,157,116,.4)}.pod.rank-1::before{content:'';position:absolute;top:0;left:0;right:0;height:3px;background:linear-gradient(90deg,var(--gold),#e8c87c,var(--gold))}
.pod.rank-2::before{content:'';position:absolute;top:0;left:0;right:0;height:2px;background:linear-gradient(90deg,var(--teal),#5ab8cc)}
.pod.rank-3::before{content:'';position:absolute;top:0;left:0;right:0;height:2px;background:linear-gradient(90deg,#9a7d50,#c8a06a)}
.pod-rank-row{display:flex;align-items:center;gap:8px;margin-bottom:10px}
.medal{width:20px;height:20px;border-radius:50%;display:grid;place-items:center;font-size:.6rem;font-weight:700}
.m1{background:linear-gradient(135deg,#c9a84c,#f0d060);color:#1f2839}.m2{background:linear-gradient(135deg,#7a9fb5,#aaccdc);color:#1f2839}.m3{background:linear-gradient(135deg,#8c6a42,#c99a5e);color:#fff}.mx{background:var(--bg2);color:var(--text3);font-size:.55rem}
.pod-pos{font-family:var(--mono);font-size:.58rem;letter-spacing:.15em;color:var(--text3)}
.pod-name{font-size:.9rem;font-weight:700;color:var(--navy);line-height:1.2;margin-bottom:2px}
.pod-code{font-family:var(--mono);font-size:.6rem;color:var(--teal);letter-spacing:.1em;margin-bottom:14px}
.pod-score-lbl{font-size:.55rem;font-weight:600;letter-spacing:.15em;text-transform:uppercase;color:var(--text3);margin-bottom:5px}
.pod-score-row{display:flex;align-items:center;gap:10px;margin-bottom:10px}
.pod-score-num{font-size:1.7rem;font-weight:700;line-height:1;min-width:52px}
.pod-bar-track{flex:1;height:4px;background:var(--bg2);border-radius:999px;overflow:hidden}
.pod-bar-fill{height:100%;border-radius:999px;transition:width 1.2s cubic-bezier(.22,1,.36,1);box-shadow:inset 0 0 6px rgba(255,255,255,.45)}
.pod-splits{display:flex;gap:12px;margin-bottom:12px}
.split{flex:1}.split-lbl{font-size:.55rem;font-weight:600;letter-spacing:.1em;text-transform:uppercase;color:var(--text3);margin-bottom:2px}
.split-val{font-family:var(--mono);font-size:.72rem;color:var(--text)}
.pod-footer{display:flex;align-items:center;justify-content:space-between;padding-top:10px;border-top:1px solid var(--border)}
.badge{display:inline-flex;align-items:center;justify-content:center;padding:2px 9px;border-radius:4px;font-family:var(--mono);font-size:.6rem;font-weight:500;letter-spacing:.06em;border:1px solid}
.b-aaa{color:var(--c-aaa);background:var(--c-aaa-bg);border-color:var(--c-aaa-bdr)}.b-aa{color:var(--c-aa);background:var(--c-aa-bg);border-color:var(--c-aa-bdr)}.b-a{color:var(--c-a);background:var(--c-a-bg);border-color:var(--c-a-bdr)}.b-bbb{color:var(--c-bbb);background:var(--c-bbb-bg);border-color:var(--c-bbb-bdr)}.b-bb{color:var(--c-bb);background:var(--c-bb-bg);border-color:var(--c-bb-bdr)}.b-b{color:var(--c-b);background:var(--c-b-bg);border-color:var(--c-b-bdr)}.b-low{color:var(--c-low);background:var(--c-low-bg);border-color:var(--c-low-bdr)}
.status-wrap{display:flex;align-items:center;gap:5px;white-space:nowrap}
.dot-s{width:6px;height:6px;border-radius:50%;position:relative;flex-shrink:0}
.dot-aprov{background:var(--teal)}.dot-anal{background:var(--gold)}.dot-reprov{background:var(--c-low)}.dot-fgc{background:#7a6518}
.dot-anal::after{content:'';position:absolute;inset:0;border-radius:50%;background:var(--gold);animation:dotPulse 1.8s ease-out infinite;pointer-events:none}
@keyframes dotPulse{0%{transform:scale(1);opacity:.7}70%{transform:scale(2.6);opacity:0}100%{transform:scale(2.6);opacity:0}}
.status-txt{font-size:.68rem;color:var(--text2);font-weight:500}
.filter-row{display:flex;gap:8px;align-items:center;flex-wrap:wrap;margin-bottom:16px;position:sticky;top:60px;z-index:20;background:rgba(244,245,240,.88);backdrop-filter:blur(10px);padding:12px 0;margin-left:-8px;margin-right:-8px;padding-left:8px;padding-right:8px;border-bottom:1px solid rgba(221,224,216,.4)}
.fbtn{background:var(--surface);border:1px solid var(--border);color:var(--text2);font-family:var(--font);font-size:.7rem;font-weight:500;padding:6px 14px;border-radius:6px;cursor:pointer;transition:all .18s}
.fbtn:hover{border-color:var(--teal);color:var(--teal)}.fbtn.active{background:var(--navy);border-color:var(--navy);color:#fff}
.search-wrap{position:relative;margin-left:auto}
.search-in{background:var(--surface);border:1px solid var(--border);color:var(--text);font-family:var(--font);font-size:.72rem;padding:7px 12px 7px 32px;border-radius:6px;outline:none;transition:border-color .18s;width:200px}
.search-in:focus{border-color:var(--teal)}.search-in::placeholder{color:var(--text3)}
.si{position:absolute;left:10px;top:50%;transform:translateY(-50%);width:12px;height:12px;color:var(--text3);pointer-events:none}
.cmp-btn{display:inline-grid;place-items:center;width:22px;height:22px;border-radius:6px;background:var(--bg2);border:1px solid var(--border);color:var(--text3);cursor:pointer;transition:all .18s;padding:0;margin-left:6px;flex-shrink:0}
.cmp-btn:hover{background:var(--teal);border-color:var(--teal);color:#fff;transform:scale(1.08)}
.cmp-btn svg{width:11px;height:11px;stroke-width:1.7}
.pod-name-row{display:flex;align-items:center;gap:4px;margin-bottom:2px}
.pod-name-row .pod-name{flex:1}
.tbl-wrap{overflow-x:auto;border:1px solid var(--border);border-radius:10px;background:var(--surface);box-shadow:0 1px 8px rgba(31,40,57,.05)}
.ratings-card{width:100%;display:flex;flex-direction:row;gap:6px;align-items:stretch;margin:10px 0 12px}
.rc-row{flex:1;min-width:0;display:flex;flex-direction:column;gap:4px;padding:10px 10px 8px;background:var(--surface);border:1px solid var(--border);border-radius:8px;cursor:pointer;transition:all .18s}
.rc-row:hover{transform:translateY(-2px);box-shadow:0 6px 16px rgba(31,40,57,.08)}
.rc-row.active{background:rgba(0,103,123,.08);border-color:rgba(0,103,123,.22)}
.rc-rating-badge{align-self:flex-start}
.rc-limit{font-family:var(--mono);font-size:.9rem;font-weight:700;line-height:1.1}
.rc-track{height:3px;background:var(--bg2);border-radius:999px;overflow:hidden}
.rc-fill{height:100%;border-radius:999px}
.rc-count{font-size:.55rem;color:var(--text3);font-family:var(--mono)}
.rc-note{font-family:var(--mono);font-size:.55rem;letter-spacing:.1em;color:var(--text3);margin-top:6px;text-align:right}
@media(max-width:900px){.ratings-card{flex-wrap:wrap}.rc-row{min-width:calc(50% - 3px)}}
@media(max-width:600px){.rc-row{min-width:100%}}
table{width:100%;border-collapse:collapse;min-width:880px}
thead{background:var(--surface2);position:sticky;top:0;z-index:15}
th{padding:11px 14px;text-align:left;font-size:.6rem;font-weight:600;letter-spacing:.15em;color:var(--text3);text-transform:uppercase;border-bottom:1px solid var(--border);white-space:nowrap;cursor:pointer;user-select:none;transition:color .15s}
th:hover,th.sorted{color:var(--teal)}
.sort-icon{margin-left:3px;opacity:.4;font-size:.55rem}
th.sorted .sort-icon{opacity:1;color:var(--gold)}
tbody tr{border-bottom:1px solid rgba(221,224,216,.6);cursor:pointer;transition:background .18s,transform .18s,box-shadow .18s;position:relative}
tbody tr td:first-child{border-left:3px solid transparent;transition:border-color .18s}
tbody tr:hover{background:rgba(0,103,123,.045);transform:scale(1.002)}
tbody tr:hover td:first-child{border-left-color:var(--teal)}
tbody tr.top3-row{background:rgba(182,157,116,.04)}
tbody tr.top3-row:hover{background:rgba(182,157,116,.08)}
td{padding:11px 14px;font-size:.78rem;vertical-align:middle}
.rank-cell{display:flex;align-items:center;gap:7px}
.rank-n{font-family:var(--mono);font-size:.68rem;color:var(--text3);min-width:18px;text-align:right}
.co-name-wrap{display:flex;align-items:center;gap:4px}
.co-name{font-weight:600;color:var(--navy);font-size:.8rem}.co-fiador{font-size:.65rem;color:var(--text3);margin-top:1px}.co-code{font-family:var(--mono);font-size:.6rem;color:var(--teal);margin-top:2px;letter-spacing:.06em}
.bar-cell{min-width:130px}
.bar-row{display:flex;align-items:center;gap:7px}
.bar-n{font-family:var(--mono);font-size:.7rem;color:var(--text);min-width:34px;text-align:right}
.bar-track{flex:1;height:5px;background:var(--bg2);border-radius:999px;overflow:hidden}
.bar-fill{height:100%;border-radius:999px;transition:width 1.3s cubic-bezier(.22,1,.36,1);box-shadow:inset 0 0 6px rgba(255,255,255,.5),inset 0 1px 0 rgba(255,255,255,.3)}
.empty{text-align:center;padding:56px;color:var(--text3);font-size:.85rem;font-weight:500}
#tbl-count,#btbl-count{font-family:var(--mono);font-size:.58rem;color:var(--text3);letter-spacing:.1em;text-align:right;margin-top:8px}
footer{position:relative;z-index:1;background:var(--navy);border-top:1px solid rgba(255,255,255,.08);padding:24px 48px;display:flex;align-items:center;justify-content:space-between}
.foot-copy{font-family:var(--mono);font-size:.55rem;color:rgba(213,216,201,.35);letter-spacing:.1em}.foot-brand{font-size:.75rem;font-weight:700;letter-spacing:.18em;color:var(--gold);text-transform:uppercase}

/* ═══ OLD DRAWER MODAL (Tabelas e Cards) ═══ */
#modal-overlay{position:fixed;inset:0;background:rgba(31,40,57,.42);z-index:500;opacity:0;pointer-events:none;transition:opacity .3s,backdrop-filter .3s;backdrop-filter:blur(4px);-webkit-backdrop-filter:blur(4px)}
#modal-overlay.open{opacity:1;pointer-events:all}
#modal{position:fixed;top:0;right:0;bottom:0;width:min(540px,96vw);background:var(--surface);z-index:501;transform:translateX(100%);transition:transform .38s cubic-bezier(.32,0,.17,1);overflow-y:auto;display:flex;flex-direction:column;box-shadow:-20px 0 60px rgba(31,40,57,.15)}
#modal.open{transform:translateX(0)}
.modal-hdr{background:var(--navy);padding:28px 28px 24px;position:sticky;top:0;z-index:10;flex-shrink:0}
.modal-close{position:absolute;top:16px;right:16px;width:32px;height:32px;border:1px solid rgba(255,255,255,.15);border-radius:6px;background:transparent;cursor:pointer;display:grid;place-items:center;color:rgba(213,216,201,.6);transition:background .15s,color .15s;font-size:1rem}
.modal-close:hover{background:rgba(255,255,255,.1);color:#fff}
.modal-hdr-top{display:flex;align-items:flex-start;gap:12px;margin-bottom:16px}
.modal-rank-badge{width:36px;height:36px;border-radius:8px;background:rgba(182,157,116,.15);border:1px solid rgba(182,157,116,.3);display:grid;place-items:center;font-size:.7rem;font-weight:700;color:var(--gold);font-family:var(--mono);flex-shrink:0}
.modal-co-name{font-size:1.2rem;font-weight:700;color:#fff;line-height:1.2;margin-bottom:2px}
.modal-co-sub{font-family:var(--mono);font-size:.62rem;color:rgba(213,216,201,.55);letter-spacing:.12em}
.modal-scores{display:grid;grid-template-columns:1fr 1fr 1fr;gap:10px}
.mscore{background:rgba(255,255,255,.05);border:1px solid rgba(255,255,255,.08);border-radius:8px;padding:12px;text-align:center}
.mscore-lbl{font-size:.55rem;font-weight:600;letter-spacing:.12em;text-transform:uppercase;color:rgba(213,216,201,.5);margin-bottom:5px}
.mscore-num{font-size:1.4rem;font-weight:700;letter-spacing:-.02em;line-height:1}
.mscore-bar{margin-top:6px;height:4px;background:rgba(255,255,255,.1);border-radius:999px;overflow:hidden}
.mscore-fill{height:100%;border-radius:999px;transition:width 1s .3s ease;box-shadow:inset 0 0 6px rgba(255,255,255,.3)}
.modal-body{flex:1;padding:24px 28px}
.modal-section{margin-bottom:28px}
.modal-sec-title{font-size:.6rem;font-weight:700;letter-spacing:.2em;text-transform:uppercase;color:var(--teal);margin-bottom:14px;display:flex;align-items:center;gap:8px}
.modal-sec-title::after{content:'';flex:1;height:1px;background:var(--border)}
.kpi-grid{display:grid;grid-template-columns:repeat(3,1fr);gap:10px}
.kpi{background:var(--surface2);border:1px solid var(--border);border-radius:8px;padding:12px;position:relative;overflow:hidden}
.kpi-lbl{font-size:.56rem;font-weight:600;letter-spacing:.1em;text-transform:uppercase;color:var(--text3);margin-bottom:5px;line-height:1.3}
.kpi-val{font-family:var(--mono);font-size:.95rem;font-weight:500;color:var(--navy)}
.kpi-val.pos{color:#0d8a7a}.kpi-val.neg{color:var(--c-low)}.kpi-val.neu{color:var(--navy)}
.kpi-context{font-size:.58rem;color:var(--text3);margin-top:2px;font-weight:500}
.kpi-bar-bg{position:absolute;bottom:0;left:0;right:0;height:3px;background:var(--border);border-radius:999px}
.kpi-bar-fg{position:absolute;bottom:0;left:0;height:3px;border-radius:999px;transition:width .8s .4s ease;box-shadow:inset 0 0 4px rgba(255,255,255,.4)}
.qual-grid{display:grid;grid-template-columns:1fr 1fr;gap:10px;margin-bottom:16px}
.qdim{background:var(--surface2);border:1px solid var(--border);border-radius:8px;padding:12px}
.qdim-lbl{font-size:.56rem;font-weight:600;letter-spacing:.1em;text-transform:uppercase;color:var(--text3);margin-bottom:8px}
.qdim-row{display:flex;align-items:center;gap:8px}
.qdim-dots{display:flex;gap:3px}
.qdim-dot{width:9px;height:9px;border-radius:2px;border:1px solid var(--border)}.qdim-dot.filled{border-color:var(--teal);background:var(--teal)}
.qdim-val{font-family:var(--mono);font-size:.7rem;font-weight:500;color:var(--text2)}
.cmt-card{background:var(--surface2);border:1px solid var(--border);border-left:3px solid var(--teal);border-radius:0 8px 8px 0;padding:12px 14px;margin-bottom:8px}
.cmt-lbl{font-size:.56rem;font-weight:700;letter-spacing:.12em;text-transform:uppercase;color:var(--teal);margin-bottom:5px}
.cmt-txt{font-size:.74rem;color:var(--text2);line-height:1.55;font-weight:400}
.setor-tag{display:inline-flex;align-items:center;gap:6px;background:linear-gradient(135deg,rgba(31,40,57,.06),rgba(0,103,123,.06));border:1px solid rgba(0,103,123,.2);border-radius:20px;padding:5px 12px 5px 8px;margin-bottom:14px}
.setor-icon{width:18px;height:18px;background:var(--navy);border-radius:50%;display:grid;place-items:center;font-size:.55rem;color:var(--offwhite);font-weight:700;flex-shrink:0}
.setor-label{font-size:.6rem;font-weight:700;letter-spacing:.08em;text-transform:uppercase;color:var(--text3)}
.setor-val{font-size:.72rem;font-weight:600;color:var(--navy)}
.meta-row{display:flex;gap:10px;flex-wrap:wrap;margin-top:10px}
.meta-chip{background:var(--bg2);border:1px solid var(--border);border-radius:6px;padding:6px 12px;font-size:.65rem;font-weight:600;color:var(--text2)}
.meta-chip span{color:var(--teal)}

/* ═══ OMNI MODAL (BENTO BOX / INSPECTOR STYLE - APENAS SEARCH) ═══ */
#omni-overlay { position:fixed; inset:0; background:rgba(15,23,42,0.6); backdrop-filter:blur(8px); -webkit-backdrop-filter:blur(8px); z-index:2000; opacity:0; pointer-events:none; transition:opacity .3s ease; display:grid; place-items:center; padding:20px; }
#omni-overlay.open { opacity:1; pointer-events:auto; }
#omni-modal { width:100%; max-width:1040px; background:var(--surface); border-radius:20px; box-shadow:0 32px 90px rgba(15,23,42,.3); overflow:hidden; transform:scale(0.96) translateY(20px); opacity:0; transition:all .4s cubic-bezier(0.16, 1, 0.3, 1); margin:auto; display:flex; flex-direction:column; max-height:88vh; }
#omni-overlay.open #omni-modal { transform:scale(1) translateY(0); opacity:1; }

.omni-app-layout { display:flex; flex-direction:column; height:100%; }
.omni-hdr { display:flex; justify-content:space-between; align-items:flex-start; padding:24px 32px; border-bottom:1px solid var(--border); background:var(--surface); flex-shrink:0; }
.omni-title-wrap { display:flex; gap:16px; align-items:center; }
.omni-rating-badge { padding:8px 16px; border-radius:10px; font-family:var(--mono); font-size:1.3rem; font-weight:700; border:2px solid; display:grid; place-items:center; }
.omni-name { font-size:1.6rem; font-weight:800; color:var(--navy); letter-spacing:-.02em; line-height:1; margin-bottom:6px; }
.omni-meta { font-family:var(--mono); font-size:.65rem; color:var(--text3); letter-spacing:.08em; text-transform:uppercase; font-weight:600;}

.omni-close { width:32px; height:32px; background:var(--surface2); border:1px solid var(--border); border-radius:50%; color:var(--text2); display:grid; place-items:center; cursor:pointer; transition:all .2s; }
.omni-close:hover { background:var(--border); color:var(--navy); transform:scale(1.05); }

.omni-body { display:flex; flex:1; overflow:hidden; }
.omni-main { flex:2; padding:32px; background:var(--bg); overflow-y:auto; }
.omni-side { flex:1; padding:32px; background:var(--surface); border-left:1px solid var(--border); overflow-y:auto; }

.omni-sec-title { font-size:.6rem; font-weight:800; letter-spacing:.15em; text-transform:uppercase; color:var(--text3); margin-bottom:16px; }

/* Scores */
.omni-scores-row { display:grid; grid-template-columns:repeat(3, 1fr); gap:16px; margin-bottom:32px; }
.omni-score-card { background:var(--surface); border:1px solid var(--border); border-radius:12px; padding:20px 16px; display:flex; flex-direction:column; align-items:center; justify-content:center; box-shadow:0 2px 8px rgba(31,40,57,.02); }
.omni-score-card.highlight { background:linear-gradient(135deg, var(--navy), #2a364f); border:none; box-shadow:0 8px 24px rgba(31,40,57,.15); }
.omni-score-lbl { font-size:.55rem; font-weight:700; letter-spacing:.1em; text-transform:uppercase; color:var(--text3); margin-bottom:8px; }
.omni-score-val { font-size:2rem; font-weight:800; font-family:var(--mono); line-height:1; }

/* KPIs Grid */
.omni-kpi-grid { display:grid; grid-template-columns:repeat(auto-fit, minmax(160px, 1fr)); gap:12px; }
.omni-kpi-card { background:var(--surface); border:1px solid var(--border); border-radius:10px; padding:16px; transition:transform .2s, border-color .2s; cursor:default; }
.omni-kpi-card:hover { transform:translateY(-2px); border-color:var(--teal); box-shadow:0 4px 12px rgba(0,103,123,.05); }
.omni-kpi-lbl { font-size:.55rem; font-weight:600; color:var(--text3); text-transform:uppercase; letter-spacing:.05em; margin-bottom:6px; line-height:1.2; }
.omni-kpi-val { font-family:var(--mono); font-size:1.1rem; font-weight:600; }
.omni-kpi-val.pos { color:#0d8a7a; } .omni-kpi-val.neg { color:var(--c-low); } .omni-kpi-val.neu { color:var(--navy); }

/* Quali List */
.omni-qual-list { display:flex; flex-direction:column; gap:16px; }
.omni-qdim { display:flex; flex-direction:column; gap:6px; cursor:pointer; padding:8px 12px; margin:0 -12px; border-radius:8px; border:1px solid transparent; transition:all .2s cubic-bezier(0.16, 1, 0.3, 1); }
.omni-qdim:hover, .omni-qdim.hovered { background:rgba(0,103,123,.04); transform:translateX(4px); }
.omni-qdim.locked { background:var(--surface2); border-color:var(--teal); padding:10px 14px; margin:0 -14px; box-shadow:0 4px 12px rgba(0,103,123,.08); transform:translateX(4px); }
.omni-qdim-top { display:flex; justify-content:space-between; align-items:center; }
.omni-qdim-lbl { font-size:.65rem; font-weight:600; color:var(--navy); transition:color .2s; }
.omni-qdim.locked .omni-qdim-lbl { color:var(--teal); }
.omni-qdim-val { font-family:var(--mono); font-size:.7rem; font-weight:700; color:var(--teal); }
.omni-qdim-track { width:100%; height:6px; background:var(--bg2); border-radius:999px; overflow:hidden; }
.omni-qdim-fill { height:100%; border-radius:999px; transition:width 1s cubic-bezier(0.16, 1, 0.3, 1); }
.omni-qdim.locked .omni-qdim-fill { box-shadow: 0 0 8px rgba(0,103,123,.4); }

/* Comments Display Area */
.omni-cmt-display { flex:1; min-height:160px; display:flex; flex-direction:column; justify-content:flex-start; overflow-y:auto; padding-right:8px; }
.omni-cmt-display::-webkit-scrollbar { width:4px; }
.omni-cmt-placeholder { text-align:center; font-size:.65rem; color:var(--text3); font-weight:500; font-family:var(--mono); border:1px dashed var(--border); padding:32px 20px; border-radius:8px; opacity:0.6; display:flex; flex-direction:column; align-items:center; gap:12px; margin-top:auto; margin-bottom:auto;}
.omni-cmt-card { background:transparent; padding:0; margin:0; }
.omni-cmt-lbl { font-size:.6rem; font-weight:800; letter-spacing:.1em; text-transform:uppercase; color:var(--teal); margin-bottom:8px; }
.omni-cmt-txt { font-size:.75rem; color:var(--text2); line-height:1.6; }
.fade-in { animation: fadeIn .3s cubic-bezier(0.16, 1, 0.3, 1) forwards; }
@keyframes fadeIn { from { opacity:0; transform:translateY(8px); } to { opacity:1; transform:translateY(0); } }

@media (max-width:900px) {
  .omni-body { flex-direction:column; overflow-y:auto; }
  .omni-main, .omni-side { overflow-y:visible; flex:none; border-left:none; border-top:1px solid var(--border); }
  .omni-cmt-display { min-height: 120px; margin-top:16px; }
}

/* ═══ COMPARADOR ═══ */
#comparador,#bcomparador{position:relative;z-index:1;max-width:1400px;margin:0 auto;padding:0 48px 56px}
.comp-shell{background:var(--surface);border:1px solid var(--border);border-radius:16px;overflow:hidden;box-shadow:0 4px 24px rgba(31,40,57,.06)}
.comp-topbar{background:var(--navy);padding:20px 28px;display:flex;align-items:center;justify-content:space-between;border-bottom:1px solid rgba(255,255,255,.06)}
.comp-topbar-left{display:flex;align-items:center;gap:14px}
.comp-terminal-dot{display:flex;gap:5px}.comp-td{width:8px;height:8px;border-radius:50%}
.comp-td-r{background:#c0392b}.comp-td-y{background:#b69d74}.comp-td-g{background:#00677b}
.comp-topbar-title{font-family:var(--mono);font-size:.68rem;letter-spacing:.2em;color:rgba(213,216,201,.6);text-transform:uppercase}
.comp-topbar-badge{font-family:var(--mono);font-size:.55rem;color:rgba(0,103,123,.7);background:rgba(0,103,123,.1);border:1px solid rgba(0,103,123,.2);border-radius:4px;padding:2px 8px;letter-spacing:.1em}
.comp-selectors{display:grid;grid-template-columns:1fr 1fr;gap:0;border-bottom:1px solid var(--border)}
.comp-sel-col{padding:20px 28px;position:relative}
.comp-sel-col:first-child{border-right:1px solid var(--border)}
.comp-sel-col.col-a{border-top:3px solid var(--teal)}
.comp-sel-col.col-b{border-top:3px solid var(--gold)}
.comp-sel-label{font-size:.58rem;font-weight:700;letter-spacing:.2em;text-transform:uppercase;margin-bottom:10px;display:flex;align-items:center;gap:8px}
.comp-sel-label.la{color:var(--teal)}.comp-sel-label.lb{color:var(--gold)}
.comp-sel-dot{width:8px;height:8px;border-radius:50%;flex-shrink:0}
.da{background:var(--teal)}.db{background:var(--gold)}
.comp-input-wrap{position:relative}
.comp-input{width:100%;background:var(--bg);border:1px solid var(--border);color:var(--text);font-family:var(--font);font-size:.8rem;font-weight:500;padding:10px 14px;border-radius:8px;outline:none;transition:border-color .18s,box-shadow .18s;cursor:text}
.comp-input:focus{border-color:var(--teal);box-shadow:0 0 0 3px rgba(0,103,123,.08)}
.comp-sel-col.col-b .comp-input:focus{border-color:var(--gold);box-shadow:0 0 0 3px rgba(182,157,116,.1)}
.comp-input::placeholder{color:var(--text3);font-weight:400;font-size:.75rem}
.comp-dropdown{position:absolute;top:calc(100% + 4px);left:0;right:0;background:var(--surface);border:1px solid var(--border);border-radius:8px;box-shadow:0 12px 32px rgba(31,40,57,.12);z-index:50;max-height:220px;overflow-y:auto;display:none}
.comp-dropdown.open{display:block}
.comp-opt{padding:9px 14px;font-size:.76rem;font-weight:500;color:var(--text);cursor:pointer;display:flex;align-items:center;justify-content:space-between;transition:background .1s}
.comp-opt:hover{background:rgba(0,103,123,.05)}
.comp-opt-rating{font-family:var(--mono);font-size:.6rem;color:var(--text3)}
.comp-sel-chosen{display:flex;align-items:center;gap:10px;margin-top:8px;padding:8px 12px;background:var(--surface2);border:1px solid var(--border);border-radius:6px;display:none}
.comp-sel-chosen.visible{display:flex}
.comp-chosen-name{font-size:.74rem;font-weight:600;color:var(--navy);flex:1}
.comp-chosen-meta{font-family:var(--mono);font-size:.58rem;color:var(--text3)}
.comp-clear{background:none;border:none;cursor:pointer;color:var(--text3);font-size:.7rem;padding:2px 4px;border-radius:3px;transition:color .15s}
.comp-clear:hover{color:var(--c-low)}
.comp-body{padding:28px;display:none;opacity:0;transition:opacity .4s}
.comp-body.visible{display:block;opacity:1}
.comp-main-grid{display:grid;grid-template-columns:1fr 1.1fr;gap:24px;align-items:start;margin-bottom:28px}
.comp-radar-wrap{background:var(--surface2);border:1px solid var(--border);border-radius:12px;padding:20px;position:relative}
.comp-radar-title{font-size:.6rem;font-weight:700;letter-spacing:.18em;text-transform:uppercase;color:var(--text3);margin-bottom:16px;display:flex;align-items:center;gap:8px}
.comp-radar-title::after{content:'';flex:1;height:1px;background:var(--border)}
.comp-radar-canvas-wrap{position:relative;height:280px}
.comp-radar-legend{display:flex;gap:20px;justify-content:center;margin-top:14px}
.comp-legend-item{display:flex;align-items:center;gap:6px;font-size:.65rem;font-weight:600;color:var(--text2)}
.comp-legend-dot{width:10px;height:10px;border-radius:2px;flex-shrink:0}
.comp-scores-panel{display:flex;flex-direction:column;gap:12px}
.comp-score-row{display:grid;grid-template-columns:1fr auto 1fr;gap:10px;align-items:center;background:var(--surface2);border:1px solid var(--border);border-radius:10px;padding:14px 16px}
.comp-score-col{display:flex;flex-direction:column;gap:3px}
.comp-score-col.right{text-align:right;align-items:flex-end}
.comp-score-lbl-tiny{font-size:.52rem;font-weight:600;letter-spacing:.14em;text-transform:uppercase;color:var(--text3)}
.comp-score-val{font-family:var(--mono);font-size:1.35rem;font-weight:700;letter-spacing:-.02em;line-height:1}
.comp-score-val.winner::after{content:'▲';font-size:.45rem;margin-left:3px;vertical-align:super;opacity:.7}
.comp-score-bar-center{display:flex;flex-direction:column;align-items:center;gap:4px}
.comp-score-bar-lbl{font-size:.5rem;font-weight:600;letter-spacing:.14em;text-transform:uppercase;color:var(--text3)}
.comp-duel-bar{width:6px;height:60px;background:var(--border);border-radius:3px;position:relative;overflow:hidden}
.comp-duel-fill-a{position:absolute;bottom:50%;left:0;right:0;background:var(--teal);border-radius:3px 3px 0 0;transition:height .9s cubic-bezier(.22,1,.36,1)}
.comp-duel-fill-b{position:absolute;top:50%;left:0;right:0;background:var(--gold);border-radius:0 0 3px 3px;transition:height .9s cubic-bezier(.22,1,.36,1)}
.comp-kpis-section-title{font-size:.6rem;font-weight:700;letter-spacing:.2em;text-transform:uppercase;color:var(--teal);margin-bottom:14px;display:flex;align-items:center;gap:8px}
.comp-kpis-section-title::after{content:'';flex:1;height:1px;background:var(--border)}
.comp-kpi-table{display:flex;flex-direction:column;gap:8px}
.comp-kpi-row{display:grid;grid-template-columns:1fr 120px 1fr;gap:10px;align-items:center;padding:10px 14px;background:var(--surface2);border:1px solid var(--border);border-radius:8px;transition:border-color .2s}
.comp-kpi-row:hover{border-color:var(--border2)}
.comp-kpi-val-a,.comp-kpi-val-b{font-family:var(--mono);font-size:.8rem;font-weight:500;color:var(--text);padding:4px 10px;border-radius:5px;transition:background .3s,color .3s}
.comp-kpi-val-a{text-align:right}.comp-kpi-val-b{text-align:left}
.comp-kpi-val-a.win{background:rgba(0,103,123,.1);color:var(--teal);font-weight:700;animation:kpiPulse .6s ease}
.comp-kpi-val-b.win{background:rgba(182,157,116,.15);color:#8a7040;font-weight:700;animation:kpiPulse .6s ease}
.comp-kpi-center{text-align:center;display:flex;flex-direction:column;align-items:center;gap:2px}
.comp-kpi-name{font-size:.58rem;font-weight:700;letter-spacing:.08em;text-transform:uppercase;color:var(--text3);line-height:1.2}
.comp-kpi-ctx{font-size:.52rem;color:var(--text3);opacity:.7}
.comp-placeholder{padding:56px 28px;text-align:center}
.comp-placeholder-icon{font-size:2.5rem;margin-bottom:14px;opacity:.15;color:var(--navy)}
.comp-placeholder-txt{font-size:.82rem;color:var(--text3);font-weight:500;line-height:1.6}
.comp-placeholder-sub{font-family:var(--mono);font-size:.6rem;color:var(--text3);margin-top:6px;opacity:.6;letter-spacing:.1em}
@keyframes kpiPulse{0%{transform:scale(1)}40%{transform:scale(1.04)}100%{transform:scale(1)}}

/* ═══ KPI SETORIAL ═══ */
.kpi-tabs{display:flex;gap:8px;margin-bottom:20px;flex-wrap:wrap}
.kpi-tab{background:var(--surface);border:1px solid var(--border);border-radius:20px;padding:6px 14px;font-size:.65rem;font-weight:600;color:var(--text2);cursor:pointer;transition:all .2s;text-transform:uppercase;letter-spacing:.05em;font-family:var(--font);outline:none}
.kpi-tab:hover{border-color:var(--teal);color:var(--teal)}
.kpi-tab.active{background:var(--teal);color:#fff;border-color:var(--teal);box-shadow:0 4px 12px rgba(0,103,123,.2)}

/* ═══ BENCHMARK TOGGLE ═══ */
.bench-toggle-wrap { display:flex; align-items:center; gap:8px; }
.bench-lbl { font-size:.55rem; font-weight:700; letter-spacing:.1em; text-transform:uppercase; color:var(--text3); }
.bench-toggle { display:inline-flex; background:var(--surface2); border:1px solid var(--border); border-radius:6px; overflow:hidden; }
.bench-btn { font-family:var(--font); font-size:.6rem; font-weight:600; padding:5px 12px; cursor:pointer; background:transparent; border:none; color:var(--text3); transition:all .2s; }
.bench-btn:hover { background:rgba(31,40,57,.05); color:var(--navy); }
.bench-btn.active { background:var(--navy); color:#fff; }

/* ═══ BANKING HIGHLIGHT CHIP ═══ */
.bank-tag{display:inline-flex;align-items:center;gap:5px;background:linear-gradient(135deg,rgba(182,157,116,.08),rgba(0,103,123,.06));border:1px solid rgba(0,103,123,.18);border-radius:6px;padding:3px 9px;font-size:.6rem;font-weight:700;letter-spacing:.08em;text-transform:uppercase;color:var(--teal);margin-bottom:10px}

.reveal{opacity:0;transform:translateY(20px);transition:opacity .7s cubic-bezier(.22,1,.36,1),transform .7s cubic-bezier(.22,1,.36,1)}
.reveal.visible{opacity:1;transform:none}
.rv1{transition-delay:.06s}.rv2{transition-delay:.13s}.rv3{transition-delay:.2s}.rv4{transition-delay:.27s}.rv5{transition-delay:.34s}
@keyframes fadeUp{from{opacity:0;transform:translateY(18px)}to{opacity:1;transform:none}}
::-webkit-scrollbar{width:5px;height:5px}::-webkit-scrollbar-track{background:var(--bg)}::-webkit-scrollbar-thumb{background:var(--border2);border-radius:3px}::-webkit-scrollbar-thumb:hover{background:var(--teal)}
@media(max-width:900px){nav,.nav-inner,#hero-privado,#hero-bancos,.section,#comparador,#bcomparador,#setorial,#bsetorial{padding-left:20px;padding-right:20px}.top5-grid{grid-template-columns:1fr 1fr}.comp-main-grid{grid-template-columns:1fr}.comp-selectors{grid-template-columns:1fr}footer{flex-direction:column;gap:8px;text-align:center;padding:20px} .nav-search-wrap { display:none; } }
@media(max-width:600px){.top5-grid{grid-template-columns:1fr}.kpi-grid{grid-template-columns:1fr 1fr}.qual-grid{grid-template-columns:1fr}.comp-kpi-row{grid-template-columns:1fr 90px 1fr}.nav-tabs{display:none}}
</style>
</head>
<body>
<div id="cur"></div><div id="cur-ring"></div><div id="progress"></div><canvas id="canvas"></canvas>

<div id="modal-overlay" onclick="closeModal()"></div>
<div id="modal">
  <div class="modal-hdr" id="modal-hdr">
    <button class="modal-close" onclick="closeModal()">✕</button>
    <div class="modal-hdr-top"><div class="modal-rank-badge" id="m-rank">#1</div><div><div class="modal-co-name" id="m-name">—</div><div class="modal-co-sub" id="m-sub">—</div></div></div>
    <div class="modal-scores">
      <div class="mscore"><div class="mscore-lbl">Score Quant.</div><div class="mscore-num" id="m-sq" style="color:var(--teal)">—</div><div class="mscore-bar"><div class="mscore-fill" id="m-sq-bar" style="background:var(--teal)"></div></div></div>
      <div class="mscore"><div class="mscore-lbl">Score Qual.</div><div class="mscore-num" id="m-squ" style="color:var(--gold)">—</div><div class="mscore-bar"><div class="mscore-fill" id="m-squ-bar" style="background:var(--gold)"></div></div></div>
      <div class="mscore"><div class="mscore-lbl">Score Total</div><div class="mscore-num" id="m-st">—</div><div class="mscore-bar"><div class="mscore-fill" id="m-st-bar"></div></div></div>
    </div>
  </div>
  <div class="modal-body" id="modal-body"></div>
</div>

<div id="omni-overlay" onclick="closeOmniModal(event)">
  <div id="omni-modal" onclick="event.stopPropagation()">
    </div>
</div>

<nav><div class="nav-inner">
  <div class="nav-brand">
    <img src="{{ logo_src }}" alt="Douro Capital" style="height:24px;width:auto;display:block;">
    <div class="nav-divider"></div>
    <span class="nav-subtitle" id="nav-context-label">Crédito Privado</span>
  </div>
  <div class="nav-tabs">
    <button class="nav-tab-btn active" id="tab-privado" onclick="switchContext('privado')">Corporativo</button>
    <button class="nav-tab-btn" id="tab-bancos" onclick="switchContext('bancos')">Bancos</button>
  </div>
  
  <div class="nav-search-wrap" id="global-search-wrap">
    <div class="nav-search-box" id="gs-box">
      <svg class="ns-icon" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><circle cx="11" cy="11" r="8"/><path d="M21 21l-4.35-4.35"/></svg>
      <input type="text" class="ns-input" id="gs-input" placeholder="Buscar empresa ou banco..." autocomplete="off">
      <div class="ns-hint"><span class="ns-key">Ctrl</span><span class="ns-key">K</span></div>
    </div>
    <div class="ns-dropdown" id="gs-dropdown"></div>
  </div>
</div></nav>

<div id="ctx-privado">

<div id="hero-privado">
  <div class="hero-label">Análise de Crédito Corporativo</div>
  <h1 class="hero-title">Scorecard de Empresas</h1>
  <div class="hero-rule"></div>
  <p style="font-size:.82rem;color:var(--text2);line-height:1.7;max-width:480px;opacity:0;animation:fadeUp .7s .55s forwards">Ranking quantitativo e qualitativo de emissores de crédito privado, ordenado por score total Douro.</p>
  <div class="stats-row">
    <div class="stat-box s-total"><div class="stat-num">{{ total_p }}</div><div class="stat-lbl">Empresas</div></div>
    <div class="stat-box s-aprov"><div class="stat-num c-teal">{{ aprovadas_p }}</div><div class="stat-lbl">Aprovadas</div></div>
    <div class="stat-box s-anal"><div class="stat-num c-gold">{{ em_analise_p }}</div><div class="stat-lbl">Em análise</div></div>
    <div class="stat-box s-reprov"><div class="stat-num c-red">{{ reprovadas_p }}</div><div class="stat-lbl">Reprovadas</div></div>
  </div>
</div>

<div class="section reveal"><div class="sec-eyebrow">Destaque</div><div class="sec-title">Top <span>5</span> Empresas — Clique para Deep Dive</div><div class="top5-grid" id="top5-grid"></div></div>

<div id="comparador" class="reveal rv2">
  <div class="sec-eyebrow">Análise Comparativa</div>
  <div class="sec-title">Comparador <span>Quali × Quant</span></div>
  <div class="comp-shell">
    <div class="comp-topbar">
      <div class="comp-topbar-left"><div class="comp-terminal-dot"><div class="comp-td comp-td-r"></div><div class="comp-td comp-td-y"></div><div class="comp-td comp-td-g"></div></div><span class="comp-topbar-title">comparador.exe</span></div>
      <span class="comp-topbar-badge">DOURO // CRÉDITO</span>
    </div>
    <div class="comp-selectors">
      <div class="comp-sel-col col-a"><div class="comp-sel-label la"><div class="comp-sel-dot da"></div>Empresa A</div><div class="comp-input-wrap"><input class="comp-input" id="inp-a" type="text" placeholder="Digite para buscar..." autocomplete="off"/><div class="comp-dropdown" id="dd-a"></div></div><div class="comp-sel-chosen" id="chosen-a"><div style="flex:1"><div class="comp-chosen-name" id="chosen-name-a">—</div><div class="comp-chosen-meta" id="chosen-meta-a">—</div></div><button class="comp-clear" onclick="clearSel('a','privado')">✕</button></div></div>
      <div class="comp-sel-col col-b"><div class="comp-sel-label lb"><div class="comp-sel-dot db"></div>Empresa B</div><div class="comp-input-wrap"><input class="comp-input" id="inp-b" type="text" placeholder="Digite para buscar..." autocomplete="off"/><div class="comp-dropdown" id="dd-b"></div></div><div class="comp-sel-chosen" id="chosen-b"><div style="flex:1"><div class="comp-chosen-name" id="chosen-name-b">—</div><div class="comp-chosen-meta" id="chosen-meta-b">—</div></div><button class="comp-clear" onclick="clearSel('b','privado')">✕</button></div></div>
    </div>
    <div class="comp-placeholder" id="comp-placeholder"><div class="comp-placeholder-icon">◎</div><div class="comp-placeholder-txt">Selecione duas empresas para iniciar a análise comparativa</div><div class="comp-placeholder-sub">// radar · tug-of-war · kpis financeiros</div></div>
    <div class="comp-body" id="comp-body">
      <div class="comp-main-grid"><div class="comp-radar-wrap"><div class="comp-radar-title">Perfil Qualitativo</div><div class="comp-radar-canvas-wrap"><canvas id="radar-chart"></canvas></div><div class="comp-radar-legend"><div class="comp-legend-item"><div class="comp-legend-dot" style="background:var(--teal)"></div><span id="legend-a">Empresa A</span></div><div class="comp-legend-item"><div class="comp-legend-dot" style="background:var(--gold)"></div><span id="legend-b">Empresa B</span></div></div></div><div class="comp-scores-panel" id="comp-scores-panel"></div></div>
      <div class="comp-kpis-section-title">Cabo de Guerra — Indicadores Financeiros</div>
      <div class="comp-kpi-table" id="comp-kpi-table"></div>
    </div>
  </div>
</div>

<div id="setorial" class="reveal rv3" style="margin-bottom:56px;max-width:1400px;margin-left:auto;margin-right:auto;padding:0 48px;">
  <div class="sec-eyebrow">Visão Macro</div>
  <div class="sec-title">Benchmarking <span>Setorial</span></div>
  <div class="comp-shell">
    <div class="comp-topbar" style="background:var(--navy);">
      <div class="comp-topbar-left"><div class="comp-terminal-dot"><div class="comp-td comp-td-r"></div><div class="comp-td comp-td-y"></div><div class="comp-td comp-td-g"></div></div><span class="comp-topbar-title">benchmarking_setor.exe</span></div>
      <span class="comp-topbar-badge" style="color:var(--gold);background:rgba(182,157,116,.1);border-color:rgba(182,157,116,.2)">DOURO // MACRO</span>
    </div>
    <div class="comp-selectors" style="grid-template-columns:1fr;border-bottom:1px solid var(--border);">
      <div class="comp-sel-col col-a" style="border-right:none;border-top:3px solid var(--gold)"><div class="comp-sel-label la" style="color:var(--gold)"><div class="comp-sel-dot da" style="background:var(--gold)"></div>Selecione o Setor</div><div class="comp-input-wrap"><select id="sector-select" class="comp-input" style="appearance:auto;cursor:pointer;background:var(--surface2)"><option value="">Selecione um setor para análise quantitativa...</option></select></div></div>
    </div>
    <div class="comp-placeholder" id="sec-placeholder"><div class="comp-placeholder-icon">◱</div><div class="comp-placeholder-txt">Selecione um setor acima para visualizar o benchmarking financeiro</div></div>
    <div class="comp-body" id="sec-body" style="display:none;opacity:0;transition:opacity .4s;">
      <div class="stats-row" style="margin-top:0;margin-bottom:24px;animation:none;opacity:1;" id="sec-stats"></div>
      <div class="comp-radar-wrap" style="margin-bottom:24px;padding-bottom:10px;">
        <div class="comp-radar-title">Comparativo Quantitativo do Setor</div>
        <div style="display:flex; justify-content:space-between; align-items:center; margin-bottom:20px; flex-wrap:wrap; gap:14px;">
          <div class="bench-toggle-wrap">
            <span class="bench-lbl">Referência:</span>
            <div class="bench-toggle" id="sec-bench-toggle">
              <button class="bench-btn active" onclick="setBenchMode(event, 'privado', 'media')">Média</button>
              <button class="bench-btn" onclick="setBenchMode(event, 'privado', 'mediana')">Mediana</button>
            </div>
          </div>
          <div class="kpi-tabs" id="sec-kpi-tabs" style="margin-bottom:0;"></div>
        </div>
        <div style="position:relative;height:320px;"><canvas id="sector-chart"></canvas></div>
      </div>
      <div class="comp-kpis-section-title">Ranking Geral do Setor</div>
      <div class="tbl-wrap" style="max-height:400px;overflow-y:auto;margin-bottom:0;"><table style="min-width:700px;"><thead style="top:0;"><tr style="background:var(--surface2)"><th># Setor</th><th>Empresa</th><th>Status</th><th>Score Quant.</th><th>Score Qual.</th><th>Score Total</th><th>Rating</th></tr></thead><tbody id="sec-tbl-body"></tbody></table></div>
    </div>
  </div>
</div>

<div class="section reveal rv4">
  <div class="sec-eyebrow">Ranking Completo</div>
  <div class="sec-title">Todas as <span>Empresas</span> — Clique na linha para analisar</div>
  <div class="filter-row" id="filter-row">
    <button class="fbtn active" data-f="all" data-ctx="privado">Todas</button>
    <button class="fbtn" data-f="Aprovado" data-ctx="privado">Aprovado</button>
    <button class="fbtn" data-f="Em análise" data-ctx="privado">Em análise</button>
    <button class="fbtn" data-f="Reprovado" data-ctx="privado">Reprovado</button>
    <div class="search-wrap"><svg class="si" viewBox="0 0 16 16" fill="none" stroke="currentColor" stroke-width="1.5"><circle cx="7" cy="7" r="4.5"/><line x1="10.5" y1="10.5" x2="14" y2="14"/></svg><input class="search-in" id="search" type="text" placeholder="Buscar empresa ou código..."/></div>
  </div>
  <div class="ratings-card"></div>
  <div class="tbl-wrap"><table id="main-tbl"><thead><tr>
    <th data-col="ranking" data-ctx="privado"># <span class="sort-icon">↕</span></th>
    <th data-col="empresa" data-ctx="privado">Empresa <span class="sort-icon">↕</span></th>
    <th data-col="codigo" data-ctx="privado">Código <span class="sort-icon">↕</span></th>
    <th data-col="status" data-ctx="privado">Status <span class="sort-icon">↕</span></th>
    <th data-col="scoreQuantitativo" data-ctx="privado">Score Quant. <span class="sort-icon">↕</span></th>
    <th data-col="scoreQualitativo" data-ctx="privado">Score Qual. <span class="sort-icon">↕</span></th>
    <th data-col="scoreTotal" data-ctx="privado">Score Total <span class="sort-icon">↕</span></th>
    <th data-col="rating" data-ctx="privado">Rating <span class="sort-icon">↕</span></th>
  </tr></thead><tbody id="tbl-body"></tbody></table></div>
  <div id="tbl-count"></div>
</div>

</div><div id="ctx-bancos" style="display:none">

<div id="hero-bancos">
  <div class="hero-label">Análise de Crédito Bancário</div>
  <h1 class="hero-title">Scorecard de Bancos</h1>
  <div class="hero-rule"></div>
  <p style="font-size:.82rem;color:var(--text2);line-height:1.7;max-width:480px;opacity:0;animation:fadeUp .7s .55s forwards">Ranking de Instituições Financeiras por Basileia, Eficiência, ROE, Margem e Governança — análise proprietária Douro.</p>
  <div class="stats-row">
    <div class="stat-box s-total"><div class="stat-num">{{ total_b }}</div><div class="stat-lbl">Bancos</div></div>
    <div class="stat-box s-aprov"><div class="stat-num c-teal">{{ aprovados_b }}</div><div class="stat-lbl">Aprovados</div></div>
    <div class="stat-box s-anal"><div class="stat-num c-gold">{{ em_analise_b }}</div><div class="stat-lbl">Em análise</div></div>
    <div class="stat-box s-reprov"><div class="stat-num c-red">{{ reprovados_b }}</div><div class="stat-lbl">Reprovados / FGC</div></div>
  </div>
</div>

<div class="section reveal"><div class="sec-eyebrow">Destaque</div><div class="sec-title">Top <span>5</span> Bancos — Clique para Deep Dive</div><div class="top5-grid" id="btop5-grid"></div></div>

<div id="bcomparador" class="reveal rv2">
  <div class="sec-eyebrow">Análise Comparativa</div>
  <div class="sec-title">Comparador <span>Bancário</span></div>
  <div class="comp-shell">
    <div class="comp-topbar">
      <div class="comp-topbar-left"><div class="comp-terminal-dot"><div class="comp-td comp-td-r"></div><div class="comp-td comp-td-y"></div><div class="comp-td comp-td-g"></div></div><span class="comp-topbar-title">bank_comparador.exe</span></div>
      <span class="comp-topbar-badge">DOURO // BANCÁRIO</span>
    </div>
    <div class="comp-selectors">
      <div class="comp-sel-col col-a"><div class="comp-sel-label la"><div class="comp-sel-dot da"></div>Banco A</div><div class="comp-input-wrap"><input class="comp-input" id="binp-a" type="text" placeholder="Digite para buscar..." autocomplete="off"/><div class="comp-dropdown" id="bdd-a"></div></div><div class="comp-sel-chosen" id="bchosen-a"><div style="flex:1"><div class="comp-chosen-name" id="bchosen-name-a">—</div><div class="comp-chosen-meta" id="bchosen-meta-a">—</div></div><button class="comp-clear" onclick="clearSel('a','bancos')">✕</button></div></div>
      <div class="comp-sel-col col-b"><div class="comp-sel-label lb"><div class="comp-sel-dot db"></div>Banco B</div><div class="comp-input-wrap"><input class="comp-input" id="binp-b" type="text" placeholder="Digite para buscar..." autocomplete="off"/><div class="comp-dropdown" id="bdd-b"></div></div><div class="comp-sel-chosen" id="bchosen-b"><div style="flex:1"><div class="comp-chosen-name" id="bchosen-name-b">—</div><div class="comp-chosen-meta" id="bchosen-meta-b">—</div></div><button class="comp-clear" onclick="clearSel('b','bancos')">✕</button></div></div>
    </div>
    <div class="comp-placeholder" id="bcomp-placeholder"><div class="comp-placeholder-icon">◎</div><div class="comp-placeholder-txt">Selecione dois bancos para análise comparativa</div><div class="comp-placeholder-sub">// radar · basileia · roe · eficiência · nim</div></div>
    <div class="comp-body" id="bcomp-body">
      <div class="comp-main-grid"><div class="comp-radar-wrap"><div class="comp-radar-title">Perfil Qualitativo Bancário</div><div class="comp-radar-canvas-wrap"><canvas id="bradar-chart"></canvas></div><div class="comp-radar-legend"><div class="comp-legend-item"><div class="comp-legend-dot" style="background:var(--teal)"></div><span id="blegend-a">Banco A</span></div><div class="comp-legend-item"><div class="comp-legend-dot" style="background:var(--gold)"></div><span id="blegend-b">Banco B</span></div></div></div><div class="comp-scores-panel" id="bcomp-scores-panel"></div></div>
      <div class="comp-kpis-section-title">Cabo de Guerra — Indicadores Bancários</div>
      <div class="comp-kpi-table" id="bcomp-kpi-table"></div>
    </div>
  </div>
</div>

<div id="bsetorial" class="reveal rv3" style="margin-bottom:56px;max-width:1400px;margin-left:auto;margin-right:auto;padding:0 48px;">
  <div class="sec-eyebrow">Visão Macro</div>
  <div class="sec-title">Benchmarking <span>por Segmento</span></div>
  <div class="comp-shell">
    <div class="comp-topbar" style="background:var(--navy);">
      <div class="comp-topbar-left"><div class="comp-terminal-dot"><div class="comp-td comp-td-r"></div><div class="comp-td comp-td-y"></div><div class="comp-td comp-td-g"></div></div><span class="comp-topbar-title">benchmarking_bancario.exe</span></div>
      <span class="comp-topbar-badge" style="color:var(--gold);background:rgba(182,157,116,.1);border-color:rgba(182,157,116,.2)">DOURO // FINANCEIRO</span>
    </div>
    <div class="comp-selectors" style="grid-template-columns:1fr;border-bottom:1px solid var(--border);">
      <div class="comp-sel-col col-a" style="border-right:none;border-top:3px solid var(--gold)"><div class="comp-sel-label la" style="color:var(--gold)"><div class="comp-sel-dot da" style="background:var(--gold)"></div>Selecione o Segmento</div><div class="comp-input-wrap"><select id="bsector-select" class="comp-input" style="appearance:auto;cursor:pointer;background:var(--surface2)"><option value="">Selecione um segmento para análise...</option></select></div></div>
    </div>
    <div class="comp-placeholder" id="bsec-placeholder"><div class="comp-placeholder-icon">◱</div><div class="comp-placeholder-txt">Selecione um segmento para visualizar o benchmarking bancário</div></div>
    <div class="comp-body" id="bsec-body" style="display:none;opacity:0;transition:opacity .4s;">
      <div class="stats-row" style="margin-top:0;margin-bottom:24px;animation:none;opacity:1;" id="bsec-stats"></div>
      <div class="comp-radar-wrap" style="margin-bottom:24px;padding-bottom:10px;">
        <div class="comp-radar-title">Comparativo por Segmento</div>
        <div style="display:flex; justify-content:space-between; align-items:center; margin-bottom:20px; flex-wrap:wrap; gap:14px;">
          <div class="bench-toggle-wrap">
            <span class="bench-lbl">Referência:</span>
            <div class="bench-toggle" id="bsec-bench-toggle">
              <button class="bench-btn active" onclick="setBenchMode(event, 'bancos', 'media')">Média</button>
              <button class="bench-btn" onclick="setBenchMode(event, 'bancos', 'mediana')">Mediana</button>
            </div>
          </div>
          <div class="kpi-tabs" id="bsec-kpi-tabs" style="margin-bottom:0;"></div>
        </div>
        <div style="position:relative;height:320px;"><canvas id="bsector-chart"></canvas></div>
      </div>
      <div class="comp-kpis-section-title">Ranking Bancário do Segmento</div>
      <div class="tbl-wrap" style="max-height:400px;overflow-y:auto;margin-bottom:0;"><table style="min-width:700px;"><thead style="top:0;"><tr style="background:var(--surface2)"><th># Seg.</th><th>Banco</th><th>Status</th><th>Score Quant.</th><th>Score Qual.</th><th>Score Total</th><th>Rating</th></tr></thead><tbody id="bsec-tbl-body"></tbody></table></div>
    </div>
  </div>
</div>

<div class="section reveal rv4">
  <div class="sec-eyebrow">Ranking Completo</div>
  <div class="sec-title">Todos os <span>Bancos</span> — Clique na linha para analisar</div>
  <div class="filter-row" id="bfilter-row">
    <button class="fbtn active" data-f="all" data-ctx="bancos">Todos</button>
    <button class="fbtn" data-f="Aprovado" data-ctx="bancos">Aprovado</button>
    <button class="fbtn" data-f="Em análise" data-ctx="bancos">Em análise</button>
    <button class="fbtn" data-f="Reprovado" data-ctx="bancos">Reprovado</button>
    <button class="fbtn" data-f="FGC" data-ctx="bancos">FGC</button>
    <div class="search-wrap"><svg class="si" viewBox="0 0 16 16" fill="none" stroke="currentColor" stroke-width="1.5"><circle cx="7" cy="7" r="4.5"/><line x1="10.5" y1="10.5" x2="14" y2="14"/></svg><input class="search-in" id="bsearch" type="text" placeholder="Buscar banco ou código..."/></div>
  </div>
  <div class="tbl-wrap"><table id="bmain-tbl"><thead><tr>
    <th data-col="ranking" data-ctx="bancos"># <span class="sort-icon">↕</span></th>
    <th data-col="empresa" data-ctx="bancos">Banco <span class="sort-icon">↕</span></th>
    <th data-col="codigo" data-ctx="bancos">Código <span class="sort-icon">↕</span></th>
    <th data-col="status" data-ctx="bancos">Status <span class="sort-icon">↕</span></th>
    <th data-col="scoreQuantitativo" data-ctx="bancos">Score Quant. <span class="sort-icon">↕</span></th>
    <th data-col="scoreQualitativo" data-ctx="bancos">Score Qual. <span class="sort-icon">↕</span></th>
    <th data-col="scoreTotal" data-ctx="bancos">Score Total <span class="sort-icon">↕</span></th>
    <th data-col="rating" data-ctx="bancos">Rating <span class="sort-icon">↕</span></th>
  </tr></thead><tbody id="btbl-body"></tbody></table></div>
  <div id="btbl-count"></div>
</div>

</div><footer>
  <span class="foot-copy">// scorecard · douro capital · gerado em {{ generated_at }}</span>
  <img src="{{ logo_src }}" alt="Douro" style="height:18px;width:auto;display:block;opacity:0.8;">
</footer>

<script>
// ══════════════════════════════════════════════════
// DATA INJECTION
// ══════════════════════════════════════════════════
const DATA_PRIVADO = {{ data_privado_json }};
const DATA_BANCOS  = {{ data_bancos_json }};

// Criação do Índice Global Unificado para a Busca Central
const GLOBAL_INDEX = [
  ...DATA_PRIVADO.map((d, idx) => ({ ...d, _idx: idx, _type: 'privado' })),
  ...DATA_BANCOS.map((d, idx) => ({ ...d, _idx: idx, _type: 'bancos' }))
];

let CURRENT_CTX = 'privado';

// ══════════════════════════════════════════════════
// GLOBAL HELPERS
// ══════════════════════════════════════════════════
const ratingCls = r => {
  const m = {AAA:'aaa',AA:'aa','AA+':'aa','AA-':'aa',A:'a','A+':'a','A-':'a',BBB:'bbb','BBB+':'bbb','BBB-':'bbb',BB:'bb','BB+':'bb','BB-':'bb',B:'b','B+':'b','B-':'b'};
  return 'b-' + (m[r] || 'low');
};
const ratingColor = r => {
  const m = {AAA:'#00677b','AA+':'#00677b',AA:'#00677b','AA-':'#0d8a7a',A:'#7a6518','A+':'#7a6518','A-':'#7a6518',BBB:'#9a6e00','BBB+':'#9a6e00','BBB-':'#9a6e00',BB:'#b54a00','BB+':'#b54a00','BB-':'#b54a00',B:'#c0392b','B+':'#c0392b','B-':'#c0392b'};
  return m[r] || '#922b21';
};
const fmtN   = (v, d=2) => v == null ? '—' : (typeof v === 'number' ? v.toFixed(d) : v);
const fmtPct = v => v == null ? '—' : (v * 100).toFixed(1) + '%';
const fmtPctDirect = v => v == null ? '—' : (v * 100).toFixed(2) + '%';
const fmtPL  = v => v == null ? '—' : 'R$ ' + (v / 1e9).toFixed(1) + 'B';
const medal  = r => {
  if (r===1) return '<div class="medal m1">1</div>';
  if (r===2) return '<div class="medal m2">2</div>';
  if (r===3) return '<div class="medal m3">3</div>';
  return `<div class="medal mx">${r}</div>`;
};
const statusHTML = (s, ctx) => {
  const cls = s === 'Aprovado' ? 'dot-aprov' : s === 'Em análise' ? 'dot-anal' : s === 'FGC' ? 'dot-fgc' : 'dot-reprov';
  return `<div class="status-wrap"><div class="dot-s ${cls}"></div><span class="status-txt">${s}</span></div>`;
};
const hexLighten = (hex, amt=0.3) => {
  const h = hex.replace('#','');
  const r = parseInt(h.slice(0,2),16), g = parseInt(h.slice(2,4),16), b = parseInt(h.slice(4,6),16);
  const mix = c => Math.round(c+(255-c)*amt);
  return `rgb(${mix(r)},${mix(g)},${mix(b)})`;
};
const stripPeriod = p => { if (!p) return null; const s = String(p).trim(); if (/^trimestral$/i.test(s)||/^trim\.?$/i.test(s)) return null; return s; };
const barHTML = (v, color) => `<div class="bar-row"><span class="bar-n">${v.toFixed(1)}</span><div class="bar-track"><div class="bar-fill" data-w="${v}" style="width:0%;background:linear-gradient(90deg,${hexLighten(color,.35)},${color})"></div></div></div>`;
const dots    = (score, max=5) => { let h='<div class="qdim-dots">'; for (let i=1;i<=max;i++) h+=`<div class="qdim-dot${i<=score?' filled':''}"></div>`; return h+'</div>'; };
const scaleIconSvg = '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-linecap="round" stroke-linejoin="round"><path d="M12 3v18M5 7h14M5 7l-2 6a3 3 0 006 0L7 7M19 7l-2 6a3 3 0 006 0l-2-6"/></svg>';
const TEAL = '#00677b', GOLD = '#b69d74';

// Funções matemáticas para os gráficos
const calcMean = arr => arr.length ? arr.reduce((a,b)=>a+b,0)/arr.length : 0;
const calcMedian = arr => {
  if (!arr.length) return 0;
  const s = [...arr].sort((a,b)=>a-b);
  const mid = Math.floor(s.length/2);
  return s.length % 2 !== 0 ? s[mid] : (s[mid-1]+s[mid])/2;
};

// Limite de Exposição (Global Search e Modal)
const getAllocLimit = (rating) => {
  if (!rating) return null;
  if (rating.startsWith('AAA')) return '2,50%';
  if (rating.startsWith('AA')) return '2,25%';
  if (rating.startsWith('A') && !rating.startsWith('AA')) return '2,00%';
  if (rating.startsWith('BBB')) return '1,75%';
  if (rating.startsWith('BB') && !rating.startsWith('BBB')) return '1,50%';
  return null;
};

// Estado da linha de Benchmark
let privBenchMode = 'media';
let bankBenchMode = 'media';

window.setBenchMode = (e, ctx, mode) => {
  if (ctx === 'privado') {
    privBenchMode = mode;
    document.querySelectorAll('#sec-bench-toggle .bench-btn').forEach(b => b.classList.remove('active'));
    e.target.classList.add('active');
    if (currentSectorData.length) renderSectorChart();
  } else {
    bankBenchMode = mode;
    document.querySelectorAll('#bsec-bench-toggle .bench-btn').forEach(b => b.classList.remove('active'));
    e.target.classList.add('active');
    if (currentBankSectorData.length) renderBankSectorChart();
  }
};

// ══════════════════════════════════════════════════
// CONTEXT SWITCH
// ══════════════════════════════════════════════════
function switchContext(ctx) {
  CURRENT_CTX = ctx;
  document.getElementById('ctx-privado').style.display = ctx === 'privado' ? '' : 'none';
  document.getElementById('ctx-bancos').style.display  = ctx === 'bancos'  ? '' : 'none';
  document.getElementById('tab-privado').classList.toggle('active', ctx === 'privado');
  document.getElementById('tab-bancos').classList.toggle('active', ctx === 'bancos');
  document.getElementById('nav-context-label').textContent = ctx === 'privado' ? 'Crédito Privado' : 'Crédito Bancário';
  window.scrollTo({top: 0, behavior: 'smooth'});
  // Reinit bars after display
  setTimeout(animateBars, 120);
}

function animateBars() {
  document.querySelectorAll('.bar-fill[data-w],.pod-bar-fill[data-w]').forEach(el => {
    if (el.dataset.w) el.style.width = el.dataset.w + '%';
  });
}

// ══════════════════════════════════════════════════
// TOP 5 — PRIVADO
// ══════════════════════════════════════════════════
document.getElementById('top5-grid').innerHTML = DATA_PRIVADO.slice(0,5).map((d,i) => {
  const c = ratingColor(d.rating), cc = i===0?'rank-1':i===1?'rank-2':i===2?'rank-3':'';
  return `<div class="pod ${cc} reveal rv${i+1}" onclick="openModal(${i},'privado')">
    <div class="pod-rank-row">${medal(i+1)}<span class="pod-pos">Top ${i+1}</span></div>
    <div class="pod-name-row"><div class="pod-name">${d.empresa}</div><button class="cmp-btn" title="Comparar" onclick="event.stopPropagation();sendToCompare(${i},'privado')">${scaleIconSvg}</button></div>
    <div class="pod-code">${d.codigo}</div>
    <div class="pod-score-lbl">Score Total</div>
    <div class="pod-score-row"><div class="pod-score-num" style="color:${c}">${d.scoreTotal.toFixed(1)}</div><div class="pod-bar-track"><div class="pod-bar-fill" data-w="${d.scoreTotal}" style="width:0%;background:linear-gradient(90deg,${hexLighten(c,.35)},${c})"></div></div></div>
    <div class="pod-splits"><div class="split"><div class="split-lbl">Quant.</div><div class="split-val">${d.scoreQuantitativo.toFixed(1)}</div></div><div class="split"><div class="split-lbl">Qual.</div><div class="split-val">${d.scoreQualitativo.toFixed(1)}</div></div></div>
    <div class="pod-footer"><span class="badge ${ratingCls(d.rating)}">${d.rating}</span>${statusHTML(d.status)}</div>
  </div>`;
}).join('');

// ══════════════════════════════════════════════════
// TOP 5 — BANCOS
// ══════════════════════════════════════════════════
document.getElementById('btop5-grid').innerHTML = DATA_BANCOS.slice(0,5).map((d,i) => {
  const c = ratingColor(d.rating), cc = i===0?'rank-1':i===1?'rank-2':i===2?'rank-3':'';
  return `<div class="pod ${cc} reveal rv${i+1}" onclick="openModal(${i},'bancos')">
    <div class="pod-rank-row">${medal(i+1)}<span class="pod-pos">Top ${i+1}</span></div>
    <div class="pod-name-row"><div class="pod-name">${d.empresa}</div><button class="cmp-btn" title="Comparar" onclick="event.stopPropagation();sendToCompare(${i},'bancos')">${scaleIconSvg}</button></div>
    <div class="pod-code">${d.codigo}</div>
    <div class="pod-score-lbl">Score Total</div>
    <div class="pod-score-row"><div class="pod-score-num" style="color:${c}">${d.scoreTotal.toFixed(1)}</div><div class="pod-bar-track"><div class="pod-bar-fill" data-w="${d.scoreTotal}" style="width:0%;background:linear-gradient(90deg,${hexLighten(c,.35)},${c})"></div></div></div>
    <div class="pod-splits"><div class="split"><div class="split-lbl">Quant.</div><div class="split-val">${d.scoreQuantitativo.toFixed(1)}</div></div><div class="split"><div class="split-lbl">Qual.</div><div class="split-val">${d.scoreQualitativo.toFixed(1)}</div></div></div>
    <div class="pod-footer"><span class="badge ${ratingCls(d.rating)}">${d.rating}</span>${statusHTML(d.status)}</div>
  </div>`;
}).join('');

// ══════════════════════════════════════════════════
// TABELA GENÉRICA
// ══════════════════════════════════════════════════
let sortState = { privado: {col:'ranking',dir:1}, bancos: {col:'ranking',dir:1} };
let filterState = { privado: {status:'all',q:''}, bancos: {status:'all',q:''} };
let privadoRatingFocus = null;

function renderRatingsCard() {
  const card = document.querySelector('.ratings-card');
  if (!card) return;

  const ratingOrder = ["AAA", "AA", "A", "BBB", "BB"];
  const ratingLimits = {
    AAA: 2.50,
    AA: 2.25,
    A: 2.00,
    BBB: 1.75,
    BB: 1.50,
  };
  const ratingStyles = {
    AAA: { color: "var(--c-aaa)", width: 100 },
    AA:  { color: "var(--c-aa)",  width: 90  },
    A:   { color: "var(--c-a)",   width: 80  },
    BBB: { color: "var(--c-bbb)", width: 70  },
    BB:  { color: "var(--c-bb)",  width: 60  },
  };
  const fmtPct2 = v => `${v.toFixed(2).replace('.', ',')}%`;
  const aprovadosPorRating = ratingOrder.reduce((acc, rating) => {
    acc[rating] = DATA_PRIVADO.filter(d => d.status === "Aprovado" && d.rating === rating).length;
    return acc;
  }, {});

  const rowsHtml = ratingOrder.map((rating) => {
    const limite = ratingLimits[rating];
    const aprov = aprovadosPorRating[rating];
    const style = ratingStyles[rating];
    const activeClass = privadoRatingFocus === rating ? 'active' : '';
    return `<div class="rc-row ${activeClass}" onclick="setPrivadoRatingFocus('${rating}')" style="border-top:3px solid ${style.color}">
      <span class="badge rc-rating-badge ${ratingCls(rating)}">${rating}</span>
      <span class="rc-limit" style="color:${style.color}">${fmtPct2(limite)}</span>
      <div class="rc-track"><div class="rc-fill" style="width:${style.width}%;background:${style.color}"></div></div>
      <span class="rc-count">${aprov} aprov.</span>
    </div>`;
  }).join('');
  card.innerHTML = rowsHtml + `<div class="rc-note">// alocação máxima por emissor</div>`;
}

function setPrivadoRatingFocus(rating) {
  const searchEl = document.getElementById('search');
  if (privadoRatingFocus === rating) {
    privadoRatingFocus = null;
    filterState.privado.q = '';
    if (searchEl) searchEl.value = '';
  } else {
    privadoRatingFocus = rating;
    filterState.privado.q = rating.toLowerCase();
    if (searchEl) searchEl.value = rating;
  }
  renderTable('privado');
}

function renderTable(ctx) {
  const DATA = ctx === 'privado' ? DATA_PRIVADO : DATA_BANCOS;
  const {col,dir} = sortState[ctx];
  const {status,q} = filterState[ctx];
  const tbodyId = ctx === 'privado' ? 'tbl-body' : 'btbl-body';
  const countId = ctx === 'privado' ? 'tbl-count' : 'btbl-count';

  if (ctx === 'privado') renderRatingsCard();

  let rows = [...DATA];
  if (status !== 'all') rows = rows.filter(d => d.status === status);
  if (q) rows = rows.filter(d =>
    d.empresa.toLowerCase().includes(q) ||
    (d.codigo||'').toString().toLowerCase().includes(q) ||
    (d.fiador||'').toLowerCase().includes(q) ||
    (d.rating||'').toString().toLowerCase().includes(q)
  );
  rows.sort((a,b) => {
    const va = a[col], vb = b[col];
    return typeof va === 'string' ? dir * va.localeCompare(vb,'pt') : dir * (va - vb);
  });

  const tbody = document.getElementById(tbodyId);
  if (!rows.length) {
    tbody.innerHTML = '<tr><td colspan="8"><div class="empty">Nenhum registro encontrado.</div></td></tr>';
    document.getElementById(countId).textContent = '';
    return;
  }
  tbody.innerHTML = rows.map(d => {
    const c = ratingColor(d.rating);
    const di = DATA.findIndex(x => x.empresa === d.empresa);
    return `<tr class="${d.ranking<=3?'top3-row':''}" onclick="openModal(${di},'${ctx}')">
      <td><div class="rank-cell">${medal(d.ranking)}</div></td>
      <td><div class="co-name-wrap"><div style="flex:1"><div class="co-name">${d.empresa}</div>${d.fiador&&d.fiador!==d.empresa?`<div class="co-fiador">Fiador: ${d.fiador}</div>`:''}</div><button class="cmp-btn" title="Comparar" onclick="event.stopPropagation();sendToCompare(${di},'${ctx}')">${scaleIconSvg}</button></div></td>
      <td><span class="co-code">${d.codigo}</span></td>
      <td>${statusHTML(d.status)}</td>
      <td class="bar-cell">${barHTML(d.scoreQuantitativo,'#00677b')}</td>
      <td class="bar-cell">${barHTML(d.scoreQualitativo,c)}</td>
      <td class="bar-cell">${barHTML(d.scoreTotal,c)}</td>
      <td><span class="badge ${ratingCls(d.rating)}">${d.rating}</span></td>
    </tr>`;
  }).join('');
  document.getElementById(countId).textContent = `// Exibindo ${rows.length} de ${DATA.length} registros`;
  setTimeout(animateBars, 60);
}

// Sort listeners — privado
document.querySelectorAll('#main-tbl th[data-col]').forEach(th => {
  th.addEventListener('click', () => {
    const col = th.dataset.col;
    if (sortState.privado.col === col) sortState.privado.dir *= -1;
    else { sortState.privado.col = col; sortState.privado.dir = 1; }
    document.querySelectorAll('#main-tbl th').forEach(t => { t.classList.remove('sorted'); t.querySelector('.sort-icon').textContent='↕'; });
    th.classList.add('sorted'); th.querySelector('.sort-icon').textContent = sortState.privado.dir===1?'↑':'↓';
    renderTable('privado');
  });
});
// Sort listeners — bancos
document.querySelectorAll('#bmain-tbl th[data-col]').forEach(th => {
  th.addEventListener('click', () => {
    const col = th.dataset.col;
    if (sortState.bancos.col === col) sortState.bancos.dir *= -1;
    else { sortState.bancos.col = col; sortState.bancos.dir = 1; }
    document.querySelectorAll('#bmain-tbl th').forEach(t => { t.classList.remove('sorted'); t.querySelector('.sort-icon').textContent='↕'; });
    th.classList.add('sorted'); th.querySelector('.sort-icon').textContent = sortState.bancos.dir===1?'↑':'↓';
    renderTable('bancos');
  });
});

// Filter row — privado
document.getElementById('filter-row').addEventListener('click', e => {
  if (!e.target.classList.contains('fbtn')) return;
  document.querySelectorAll('#filter-row .fbtn').forEach(b => b.classList.remove('active'));
  e.target.classList.add('active');
  filterState.privado.status = e.target.dataset.f;
  renderTable('privado');
});
// Filter row — bancos
document.getElementById('bfilter-row').addEventListener('click', e => {
  if (!e.target.classList.contains('fbtn')) return;
  document.querySelectorAll('#bfilter-row .fbtn').forEach(b => b.classList.remove('active'));
  e.target.classList.add('active');
  filterState.bancos.status = e.target.dataset.f;
  renderTable('bancos');
});

document.getElementById('search').addEventListener('input',  e => { filterState.privado.q = e.target.value.toLowerCase().trim(); renderTable('privado'); });
document.getElementById('bsearch').addEventListener('input', e => { filterState.bancos.q  = e.target.value.toLowerCase().trim(); renderTable('bancos'); });

// ══════════════════════════════════════════════════
// GLOBAL SEARCH & COMMAND PALETTE
// ══════════════════════════════════════════════════
const gsInput = document.getElementById('gs-input');
const gsDropdown = document.getElementById('gs-dropdown');

document.addEventListener('keydown', e => {
  if ((e.ctrlKey || e.metaKey) && e.key.toLowerCase() === 'k') {
    e.preventDefault(); gsInput.focus();
  } else if (e.key === '/' && document.activeElement.tagName !== 'INPUT') {
    e.preventDefault(); gsInput.focus();
  } else if (e.key === 'Escape') {
    gsDropdown.classList.remove('open');
    closeModal(e);
    closeOmniModal(e);
  }
});

gsInput.addEventListener('input', () => {
  const q = gsInput.value.toLowerCase().trim();
  if (!q) { gsDropdown.classList.remove('open'); gsDropdown.innerHTML=''; return; }
  
  const matches = GLOBAL_INDEX.filter(d => 
    d.empresa.toLowerCase().includes(q) || 
    (d.codigo||'').toLowerCase().includes(q)
  ).slice(0, 10);
  
  if (!matches.length) { gsDropdown.classList.remove('open'); return; }
  
  gsDropdown.innerHTML = matches.map(d => {
    const isCorp = d._type === 'privado';
    const stCls = d.status === 'Aprovado' ? 'dot-aprov' : d.status === 'Em análise' ? 'dot-anal' : d.status === 'FGC' ? 'dot-fgc' : 'dot-reprov';
    
    let statusHtml = `<div class="dot-s ${stCls}"></div> <span style="color:var(--text2)">${d.status}</span>`;
    
    if (isCorp && d.status === 'Aprovado') {
      const limit = getAllocLimit(d.rating);
      if (limit) {
        statusHtml = `<div class="dot-s ${stCls}"></div> <span style="color:var(--text2)">Aprovado <span style="opacity:0.4; font-weight:400; margin:0 4px">|</span> máx ${limit} exp.</span>`;
      }
    }
    
    return `<div class="ns-item" onclick="openOmniModal(${d._idx}, '${d._type}'); closeGlobalSearch();">
      <div class="ns-item-main">
        <div class="ns-item-name">${d.empresa}</div>
        <div class="ns-item-meta">
          <span class="ns-tag ${isCorp ? 'ns-tag-corp' : 'ns-tag-banco'}">${isCorp ? 'Corp' : 'Banco'}</span>
          <span style="font-weight:600">${d.codigo}</span>
          <span class="ns-status-wrap" style="margin-left:6px">${statusHtml}</span>
        </div>
      </div>
      <div><span class="badge ${ratingCls(d.rating)}">${d.rating}</span></div>
    </div>`;
  }).join('');
  gsDropdown.classList.add('open');
});

gsInput.addEventListener('blur', () => setTimeout(() => gsDropdown.classList.remove('open'), 200));

function closeGlobalSearch() {
  gsDropdown.classList.remove('open');
  gsInput.value = '';
}

// ══════════════════════════════════════════════════
// MODAL — DRAWER LATERAL (Usado pelas Tabelas e Top 5)
// ══════════════════════════════════════════════════
function openModalPrivado(idx) {
  const d = DATA_PRIVADO[idx]; const q = d.quant||{}, ql = d.qual||{}; const c = ratingColor(d.rating);
  document.getElementById('m-rank').textContent = '#' + d.ranking;
  document.getElementById('m-name').textContent = d.empresa;
  const _periodDisp = stripPeriod(d.periodicidade);
  document.getElementById('m-sub').textContent = d.codigo + (_periodDisp?' · '+_periodDisp:'');
  document.getElementById('m-sq').textContent  = d.scoreQuantitativo.toFixed(1);  document.getElementById('m-sq-bar').style.width  = d.scoreQuantitativo+'%';
  document.getElementById('m-squ').textContent = d.scoreQualitativo.toFixed(1);   document.getElementById('m-squ-bar').style.width = d.scoreQualitativo+'%';
  document.getElementById('m-st').textContent  = d.scoreTotal.toFixed(1); document.getElementById('m-st').style.color = c;
  document.getElementById('m-st-bar').style.background = c; document.getElementById('m-st-bar').style.width = d.scoreTotal+'%';

  const kpis = [
    {lbl:'Dív.Líq./EBITDA',val:fmtN(q.div_liq_ebitda),ctx:'Alavancagem financeira',cls:q.div_liq_ebitda!=null?(q.div_liq_ebitda<3.5?'pos':'neg'):'neu',w:q.div_liq_ebitda!=null?Math.min(Math.abs(q.div_liq_ebitda)/10*100,100):0,bc:'#0d8a7a'},
    {lbl:'EBITDA/Desp.Fin.',val:fmtN(q.ebitda_desp_fin),ctx:'Cobertura de despesa financeira',cls:q.ebitda_desp_fin!=null?(q.ebitda_desp_fin>=2?'pos':'neg'):'neu',w:q.ebitda_desp_fin!=null?Math.min(q.ebitda_desp_fin/10*100,100):0,bc:'#0d8a7a'},
    {lbl:'Estrutura de Capital',val:q.estrutura_capital!=null?fmtPct(q.estrutura_capital):'—',ctx:'Dív./(Dív.+PL)',cls:'neu',w:q.estrutura_capital!=null?q.estrutura_capital*100:0,bc:q.estrutura_capital>0.7?'#c0392b':'#00677b'},
    {lbl:'Dív. CP (N)',val:q.div_cp_n!=null?'R$'+q.div_cp_n.toLocaleString('pt-BR')+'M':'—',ctx:'Curto prazo atual',cls:'neu',w:0,bc:''},
    {lbl:'Dív. CP (N-1)',val:q.div_cp_n1!=null?'R$'+q.div_cp_n1.toLocaleString('pt-BR')+'M':'—',ctx:'Período anterior',cls:'neu',w:0,bc:''},
    {lbl:'Evol. Dív. CP',val:q.evol_div_cp!=null?fmtPct(q.evol_div_cp):'—',ctx:'Variação YoY',cls:q.evol_div_cp!=null?(q.evol_div_cp<0?'pos':'neg'):'neu',w:0,bc:''},
    {lbl:'Liquidez Corrente',val:fmtN(q.liquidez_corrente),ctx:'Capacidade de pagamento de curto prazo',cls:q.liquidez_corrente!=null?(q.liquidez_corrente>=1.2?'pos':q.liquidez_corrente>=1?'neu':'neg'):'neu',w:q.liquidez_corrente!=null?Math.min(q.liquidez_corrente/3*100,100):0,bc:'#0d8a7a'},
    {lbl:'Dív.CP/Dív.Total',val:q.div_cp_div_total!=null?fmtPct(q.div_cp_div_total):'—',ctx:'Concentração CP',cls:q.div_cp_div_total!=null?(q.div_cp_div_total<0.25?'pos':'neg'):'neu',w:q.div_cp_div_total!=null?q.div_cp_div_total*100:0,bc:'#c0392b'},
    {lbl:'Margem Líquida',val:q.margem_liquida!=null?fmtPct(q.margem_liquida):'—',ctx:'Lucratividade',cls:q.margem_liquida!=null?(q.margem_liquida>0.05?'pos':q.margem_liquida>0?'neu':'neg'):'neu',w:q.margem_liquida!=null?Math.max(0,Math.min(q.margem_liquida*200,100)):0,bc:'#0d8a7a'},
    {lbl:'ROIC',val:q.roic!=null?fmtPct(q.roic):'—',ctx:'Retorno capital',cls:q.roic!=null?(q.roic>0.08?'pos':q.roic>0?'neu':'neg'):'neu',w:q.roic!=null?Math.max(0,Math.min(q.roic*400,100)):0,bc:'#0d8a7a'},
    {lbl:'Disponibilidades',val:q.disponibilidades!=null?'R$'+q.disponibilidades.toLocaleString('pt-BR')+'M':'—',ctx:'Caixa disponível',cls:'pos',w:0,bc:''},
    {lbl:'Caixa/Dív.CP',val:fmtN(q.caixa_div_cp),ctx:'Cobertura CP',cls:q.caixa_div_cp!=null?(q.caixa_div_cp>=1.5?'pos':q.caixa_div_cp>=1?'neu':'neg'):'neu',w:q.caixa_div_cp!=null?Math.min(q.caixa_div_cp/5*100,100):0,bc:'#0d8a7a'}
  ];
  const qualDims = [
    {lbl:'Vantagens Comparativas',s:ql.vantagens_score||0},
    {lbl:'Oportunidades',s:ql.oportunidades_score||0},
    {lbl:'Vulnerabilidades',s:ql.vulnerabilidades_score||0},
    {lbl:'Governança Corporativa',s:ql.governanca_score||0},
    {lbl:'Setor',s:ql.setor_score||0},
    {lbl:'Rating',s:ql.rating_score||0}
  ];
  const comments = [{lbl:'Vantagens Comparativas',txt:ql.cmt_vantagens},{lbl:'Oportunidades',txt:ql.cmt_oportunidades},{lbl:'Vulnerabilidades / Riscos',txt:ql.cmt_vulnerabilidades},{lbl:'Governança',txt:ql.cmt_governanca}].filter(c=>c.txt);
  const setor = ql.setor || null;
  const sectorHTML = setor ? `<div class="setor-tag"><div class="setor-icon">◈</div><div><div class="setor-label">Setor</div><div class="setor-val">${setor}</div></div></div>` : '';
  const _periodDisp2 = stripPeriod(d.periodicidade);
  document.getElementById('modal-body').innerHTML = `
    <div class="modal-section"><div class="modal-sec-title">Informações Gerais</div>${sectorHTML}<div class="meta-row"><div class="meta-chip"><span>Status</span> ${d.status}</div><div class="meta-chip"><span>Rating</span> ${d.rating}</div>${_periodDisp2?`<div class="meta-chip"><span>Period.</span> ${_periodDisp2}</div>`:''}</div></div>
    <div class="modal-section"><div class="modal-sec-title">Indicadores Financeiros Quantitativos</div><div class="kpi-grid">${kpis.map(k=>`<div class="kpi"><div class="kpi-lbl">${k.lbl}</div><div class="kpi-val ${k.cls}">${k.val}</div><div class="kpi-context">${k.ctx}</div><div class="kpi-bar-bg"></div>${k.w?`<div class="kpi-bar-fg" data-w="${k.w}" style="background:${k.bc};width:0%"></div>`:''}</div>`).join('')}</div></div>
    <div class="modal-section"><div class="modal-sec-title">Análise Qualitativa</div><div class="qual-grid">${qualDims.map(q=>`<div class="qdim"><div class="qdim-lbl">${q.lbl}</div><div class="qdim-row">${dots(q.s)}<span class="qdim-val">${q.s}/5</span></div></div>`).join('')}</div>${comments.map(cm=>`<div class="cmt-card"><div class="cmt-lbl">${cm.lbl}</div><div class="cmt-txt">${cm.txt}</div></div>`).join('')}</div>`;
  
  document.getElementById('modal-overlay').classList.add('open');
  document.getElementById('modal').classList.add('open');
  document.body.style.overflow = 'hidden';
  
  setTimeout(() => { document.querySelectorAll('.kpi-bar-fg[data-w]').forEach(el => el.style.width = el.dataset.w+'%'); }, 80);
}

function openModalBanco(idx) {
  const d = DATA_BANCOS[idx]; const q = d.quant||{}, ql = d.qual||{}; const c = ratingColor(d.rating);
  document.getElementById('m-rank').textContent = '#' + d.ranking;
  document.getElementById('m-name').textContent = d.empresa;
  document.getElementById('m-sub').textContent  = d.codigo + ' · Crédito Bancário';
  document.getElementById('m-sq').textContent   = d.scoreQuantitativo.toFixed(1); document.getElementById('m-sq-bar').style.width  = d.scoreQuantitativo+'%';
  document.getElementById('m-squ').textContent  = d.scoreQualitativo.toFixed(1);  document.getElementById('m-squ-bar').style.width = d.scoreQualitativo+'%';
  document.getElementById('m-st').textContent   = d.scoreTotal.toFixed(1); document.getElementById('m-st').style.color = c;
  document.getElementById('m-st-bar').style.background = c; document.getElementById('m-st-bar').style.width = d.scoreTotal+'%';

  const kpis = [
    {lbl:'Índice de Basileia',val:q.basileia!=null?fmtPctDirect(q.basileia):'—',ctx:'Suficiência de capital regulatório',cls:q.basileia!=null?(q.basileia>=0.15?'pos':q.basileia>=0.12?'neu':'neg'):'neu',w:q.basileia!=null?Math.min(q.basileia/0.25*100,100):0,bc:'#0d8a7a'},
    {lbl:'Patr. Líquido (PL)',val:fmtPL(q.pl),ctx:'Tamanho e solvência',cls:'neu',w:q.pl!=null?Math.min(q.pl/100e9*100,100):0,bc:'#00677b'},
    {lbl:'Índice de Alavancagem',val:q.alavancagem!=null?fmtPctDirect(q.alavancagem):'—',ctx:'Capital principal sobre exposição total',cls:q.alavancagem!=null?(q.alavancagem>=0.10?'pos':q.alavancagem>=0.05?'neu':'neg'):'neu',w:q.alavancagem!=null?Math.min(q.alavancagem/0.25*100,100):0,bc:'#0d8a7a'},
    {lbl:'Índice de Imobilização',val:q.imobilizacao!=null?fmtPctDirect(q.imobilizacao):'—',ctx:'Parcela do capital aplicada em ativos permanentes',cls:q.imobilizacao!=null?(q.imobilizacao<=0.20?'pos':q.imobilizacao<=0.35?'neu':'neg'):'neu',w:q.imobilizacao!=null?Math.min(q.imobilizacao/0.5*100,100):0,bc:q.imobilizacao>0.35?'#c0392b':'#0d8a7a'},
    {lbl:'Provisão/Carteira',val:q.provisao!=null?fmtPctDirect(q.provisao):'—',ctx:'Despesa de provisão sobre carteira de crédito',cls:q.provisao!=null?(q.provisao<=0.02?'pos':q.provisao<=0.05?'neu':'neg'):'neu',w:q.provisao!=null?Math.min(Math.abs(q.provisao)/0.10*100,100):0,bc:q.provisao>0.05?'#c0392b':'#0d8a7a'},
    {lbl:'Índice de Eficiência',val:q.eficiencia!=null?fmtPctDirect(q.eficiencia):'—',ctx:'Relação entre despesas operacionais e receitas',cls:q.eficiencia!=null?(q.eficiencia<=0.40?'pos':q.eficiencia<=0.60?'neu':'neg'):'neu',w:q.eficiencia!=null?Math.min(Math.abs(q.eficiencia)/2.0*100,100):0,bc:q.eficiencia>0.60?'#c0392b':'#0d8a7a'},
    {lbl:'Margem Líquida (NIM)',val:q.margem!=null?fmtPctDirect(q.margem):'—',ctx:'Rentabilidade líquida sobre operações',cls:q.margem!=null?(q.margem>=0.10?'pos':q.margem>0?'neu':'neg'):'neu',w:q.margem!=null?Math.max(0,Math.min(q.margem/0.25*100,100)):0,bc:'#0d8a7a'},
    {lbl:'ROE',val:q.roe!=null?fmtPctDirect(q.roe):'—',ctx:'Retorno sobre patrimônio líquido',cls:q.roe!=null?(q.roe>=0.15?'pos':q.roe>0?'neu':'neg'):'neu',w:q.roe!=null?Math.max(0,Math.min(q.roe/0.25*100,100)):0,bc:'#0d8a7a'},
  ];

  const qualDims = [
    {lbl:'Vantagens Comparativas', s: ql.vantagens_score||0},
    {lbl:'Oportunidades',          s: ql.oportunidades_score||0},
    {lbl:'Vulnerabilidades',       s: ql.vulnerabilidades_score||0},
    {lbl:'Governança',             s: ql.governanca_score||0},
    {lbl:'Rating Agência',         s: ql.rating_score||0},
  ];

  const comments = [
    {lbl:'Vantagens Comparativas', txt: ql.cmt_vantagens},
    {lbl:'Oportunidades',          txt: ql.cmt_oportunidades},
    {lbl:'Vulnerabilidades',       txt: ql.cmt_vulnerabilidades},
    {lbl:'Governança',             txt: ql.cmt_governanca},
    {lbl:'Rating de Agência',      txt: ql.cmt_rating},
  ].filter(c=>c.txt);

  const setor = d.setor || ql.setor || null;
  const sectorHTML = setor ? `<div class="setor-tag"><div class="setor-icon">🏦</div><div><div class="setor-label">Segmento</div><div class="setor-val">${setor}</div></div></div>` : '';

  document.getElementById('modal-body').innerHTML = `
    <div class="modal-section"><div class="modal-sec-title">Informações Gerais</div>${sectorHTML}<div class="meta-row"><div class="meta-chip"><span>Status</span> ${d.status}</div><div class="meta-chip"><span>Rating Douro</span> ${d.rating}</div>${ql.rating_txt?`<div class="meta-chip"><span>Rating Agência</span> ${ql.rating_txt}</div>`:''}</div></div>
    <div class="modal-section"><div class="modal-sec-title">Indicadores Bancários Quantitativos</div><div class="kpi-grid">${kpis.map(k=>`<div class="kpi"><div class="kpi-lbl">${k.lbl}</div><div class="kpi-val ${k.cls}">${k.val}</div><div class="kpi-context">${k.ctx}</div><div class="kpi-bar-bg"></div>${k.w?`<div class="kpi-bar-fg" data-w="${k.w}" style="background:${k.bc};width:0%"></div>`:''}</div>`).join('')}</div></div>
    <div class="modal-section"><div class="modal-sec-title">Análise Qualitativa</div><div class="qual-grid">${qualDims.map(q=>`<div class="qdim"><div class="qdim-lbl">${q.lbl}</div><div class="qdim-row">${dots(q.s)}<span class="qdim-val">${q.s}/5</span></div></div>`).join('')}</div>${comments.map(cm=>`<div class="cmt-card"><div class="cmt-lbl">${cm.lbl}</div><div class="cmt-txt">${cm.txt}</div></div>`).join('')}</div>`;

  document.getElementById('modal-overlay').classList.add('open');
  document.getElementById('modal').classList.add('open');
  document.body.style.overflow = 'hidden';
  
  setTimeout(() => { document.querySelectorAll('.kpi-bar-fg[data-w]').forEach(el => el.style.width = el.dataset.w+'%'); }, 80);
}

function openModal(idx, ctx) {
  if (ctx === 'bancos') openModalBanco(idx);
  else openModalPrivado(idx);
}

function closeModal() {
  document.getElementById('modal-overlay').classList.remove('open');
  document.getElementById('modal').classList.remove('open');
  document.body.style.overflow = '';
}


// ══════════════════════════════════════════════════
// OMNI MODAL CENTRAL (Bento Box / Inspector Style)
// ══════════════════════════════════════════════════

let lockedQdim = null;

window.hoverQdim = (el) => {
  document.querySelectorAll('.omni-qdim').forEach(q => q.classList.remove('hovered'));
  el.classList.add('hovered');
  renderCmt(el.dataset.title, el.dataset.txt);
};

window.leaveQdim = () => {
  document.querySelectorAll('.omni-qdim').forEach(q => q.classList.remove('hovered'));
  if (lockedQdim) {
    renderCmt(lockedQdim.dataset.title, lockedQdim.dataset.txt);
  } else {
    renderCmtPlaceholder();
  }
};

window.clickQdim = (el) => {
  if (lockedQdim === el) {
    // Unlock if clicking the already locked item
    lockedQdim = null;
    el.classList.remove('locked');
  } else {
    if (lockedQdim) lockedQdim.classList.remove('locked');
    lockedQdim = el;
    el.classList.add('locked');
    renderCmt(el.dataset.title, el.dataset.txt);
  }
};

function renderCmt(title, txt) {
  const display = document.getElementById('omni-cmt-display');
  if (!display) return;
  if (!txt) {
    display.innerHTML = `<div class="omni-cmt-placeholder" style="border:none">Nenhum comentário adicional para <strong>${title}</strong>.</div>`;
    return;
  }
  // Remove and re-add class to trigger CSS animation
  display.innerHTML = `
    <div class="omni-cmt-card fade-in">
      <div class="omni-cmt-lbl">${title}</div>
      <div class="omni-cmt-txt">${txt}</div>
    </div>
  `;
}

function renderCmtPlaceholder() {
  const display = document.getElementById('omni-cmt-display');
  if (!display) return;
  display.innerHTML = `
    <div class="omni-cmt-placeholder">
      <svg width="24" height="24" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.5" style="opacity:0.5; margin-bottom:4px"><path d="M15 14.5c-.9.6-2.1 1-3.5 1s-2.6-.4-3.5-1m7.5-6.5h.01M8.5 8h.01M22 12c0 5.5-4.5 10-10 10S2 17.5 2 12 6.5 2 12 2s10 4.5 10 10z"/></svg>
      Passe o cursor ou clique em um indicador acima para exibir os detalhes da análise qualitativa.
    </div>
  `;
}

function renderOmniModal(d, ctx, kpis, qualDims) {
  lockedQdim = null; // reset state
  
  const c = ratingColor(d.rating);
  const isCorp = ctx === 'privado';
  const tagType = isCorp ? 'Crédito Corporativo' : 'Crédito Bancário';
  
  let allocTag = '';
  if (isCorp && d.status === 'Aprovado') {
    const limit = getAllocLimit(d.rating);
    if (limit) allocTag = `<span style="color:var(--navy); background:rgba(0,103,123,.1); padding:2px 8px; border-radius:4px; border:1px solid rgba(0,103,123,.2)">Alocação Máx: ${limit}</span>`;
  }
  
  let html = `
    <div class="omni-app-layout">
      <div class="omni-hdr">
        <div class="omni-title-wrap">
          <div class="omni-rating-badge" style="color:${c}; background:${hexLighten(c, 0.85)}; border-color:${c}">${d.rating}</div>
          <div>
            <h2 class="omni-name">${d.empresa}</h2>
            <div class="omni-meta">${d.codigo} &nbsp;·&nbsp; Rank #${d.ranking} &nbsp;·&nbsp; ${allocTag}</div>
          </div>
        </div>
        <button class="omni-close" onclick="closeOmniModal()">
          <svg width="18" height="18" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.5"><path d="M18 6L6 18M6 6l12 12"/></svg>
        </button>
      </div>
      
      <div class="omni-body">
        <div class="omni-main">
          <div class="omni-scores-row">
            <div class="omni-score-card">
              <div class="omni-score-lbl">Score Quant.</div>
              <div class="omni-score-val" style="color:var(--teal)">${d.scoreQuantitativo.toFixed(1)}</div>
            </div>
            <div class="omni-score-card">
              <div class="omni-score-lbl">Score Qual.</div>
              <div class="omni-score-val" style="color:var(--gold)">${d.scoreQualitativo.toFixed(1)}</div>
            </div>
            <div class="omni-score-card highlight">
              <div class="omni-score-lbl">Score Total</div>
              <div class="omni-score-val" style="color:#fff">${d.scoreTotal.toFixed(1)}</div>
            </div>
          </div>
          
          <div class="omni-sec-title">Métricas Financeiras</div>
          <div class="omni-kpi-grid">
            ${kpis.map(k => `
              <div class="omni-kpi-card" title="${k.ctx}">
                <div class="omni-kpi-lbl">${k.lbl}</div>
                <div class="omni-kpi-val ${k.cls}">${k.val}</div>
              </div>
            `).join('')}
          </div>
        </div>
        
        <div class="omni-side">
          <div class="omni-sec-title">Avaliação Qualitativa</div>
          <div class="omni-qual-list">
            ${qualDims.map((q, idx) => `
              <div class="omni-qdim interactive" id="qdim-${idx}" data-idx="${idx}" data-title="${q.lbl}" data-txt="${q.txt ? q.txt.replace(/"/g, '&quot;').replace(/'/g, '&#39;') : ''}" onmouseenter="hoverQdim(this)" onmouseleave="leaveQdim()" onclick="clickQdim(this)">
                <div class="omni-qdim-top">
                  <span class="omni-qdim-lbl">${q.lbl}</span>
                  <span class="omni-qdim-val">${q.s}/5</span>
                </div>
                <div class="omni-qdim-track">
                  <div class="omni-qdim-fill" style="width:0%; background:var(--teal)" data-w="${(q.s/5)*100}"></div>
                </div>
              </div>
            `).join('')}
          </div>
          
          <div class="omni-sec-title" style="margin-top:28px">Notas do Analista</div>
          <div class="omni-cmt-display" id="omni-cmt-display">
             </div>
        </div>
      </div>
    </div>
  `;
  
  const modalEl = document.getElementById('omni-modal');
  modalEl.innerHTML = html;
  
  document.getElementById('omni-overlay').classList.add('open');
  document.body.style.overflow = 'hidden';
  
  renderCmtPlaceholder();
  
  // Animação leve das barras qualitativas após o modal abrir
  setTimeout(() => {
    document.querySelectorAll('.omni-qdim-fill[data-w]').forEach(el => {
      el.style.width = el.dataset.w + '%';
    });
  }, 150);
}

function openOmniModal(idx, ctx) {
  if (ctx === 'bancos') {
    const d = DATA_BANCOS[idx]; const q = d.quant||{}, ql = d.qual||{};
    const kpis = [
      {lbl:'Índice de Basileia',val:q.basileia!=null?fmtPctDirect(q.basileia):'—',ctx:'Suficiência de capital regulatório',cls:q.basileia!=null?(q.basileia>=0.15?'pos':q.basileia>=0.12?'neu':'neg'):'neu'},
      {lbl:'Patr. Líquido (PL)',val:fmtPL(q.pl),ctx:'Tamanho e solvência',cls:'neu'},
      {lbl:'Índice de Alavancagem',val:q.alavancagem!=null?fmtPctDirect(q.alavancagem):'—',ctx:'Capital principal sobre exposição total',cls:q.alavancagem!=null?(q.alavancagem>=0.10?'pos':q.alavancagem>=0.05?'neu':'neg'):'neu'},
      {lbl:'Índice de Imobilização',val:q.imobilizacao!=null?fmtPctDirect(q.imobilizacao):'—',ctx:'Parcela do capital aplicada em ativos permanentes',cls:q.imobilizacao!=null?(q.imobilizacao<=0.20?'pos':q.imobilizacao<=0.35?'neu':'neg'):'neu'},
      {lbl:'Provisão/Carteira',val:q.provisao!=null?fmtPctDirect(q.provisao):'—',ctx:'Despesa de provisão sobre carteira de crédito',cls:q.provisao!=null?(q.provisao<=0.02?'pos':q.provisao<=0.05?'neu':'neg'):'neu'},
      {lbl:'Índice de Eficiência',val:q.eficiencia!=null?fmtPctDirect(q.eficiencia):'—',ctx:'Relação entre despesas operacionais e receitas',cls:q.eficiencia!=null?(q.eficiencia<=0.40?'pos':q.eficiencia<=0.60?'neu':'neg'):'neu'},
      {lbl:'Margem Líquida (NIM)',val:q.margem!=null?fmtPctDirect(q.margem):'—',ctx:'Rentabilidade líquida sobre operações',cls:q.margem!=null?(q.margem>=0.10?'pos':q.margem>0?'neu':'neg'):'neu'},
      {lbl:'ROE',val:q.roe!=null?fmtPctDirect(q.roe):'—',ctx:'Retorno sobre patrimônio líquido',cls:q.roe!=null?(q.roe>=0.15?'pos':q.roe>0?'neu':'neg'):'neu'},
    ];
    const qualDims = [
      {lbl:'Vantagens Comp.', s: ql.vantagens_score||0, txt: ql.cmt_vantagens},
      {lbl:'Oportunidades',   s: ql.oportunidades_score||0, txt: ql.cmt_oportunidades},
      {lbl:'Vulnerabilidades',s: ql.vulnerabilidades_score||0, txt: ql.cmt_vulnerabilidades},
      {lbl:'Governança',      s: ql.governanca_score||0, txt: ql.cmt_governanca},
      {lbl:'Rating Agência',  s: ql.rating_score||0, txt: ql.cmt_rating},
    ];
    
    renderOmniModal(d, ctx, kpis, qualDims);
    
  } else {
    const d = DATA_PRIVADO[idx]; const q = d.quant||{}, ql = d.qual||{};
    const kpis = [
      {lbl:'Dív.Líq./EBITDA',val:fmtN(q.div_liq_ebitda),ctx:'Alavancagem financeira',cls:q.div_liq_ebitda!=null?(q.div_liq_ebitda<3.5?'pos':'neg'):'neu'},
      {lbl:'EBITDA/Desp.Fin.',val:fmtN(q.ebitda_desp_fin),ctx:'Cobertura de despesa financeira',cls:q.ebitda_desp_fin!=null?(q.ebitda_desp_fin>=2?'pos':'neg'):'neu'},
      {lbl:'Estrutura de Capital',val:q.estrutura_capital!=null?fmtPct(q.estrutura_capital):'—',ctx:'Dív./(Dív.+PL)',cls:'neu'},
      {lbl:'Liquidez Corrente',val:fmtN(q.liquidez_corrente),ctx:'Capacidade de pagamento de curto prazo',cls:q.liquidez_corrente!=null?(q.liquidez_corrente>=1.2?'pos':q.liquidez_corrente>=1?'neu':'neg'):'neu'},
      {lbl:'Dív.CP/Dív.Total',val:q.div_cp_div_total!=null?fmtPct(q.div_cp_div_total):'—',ctx:'Concentração CP',cls:q.div_cp_div_total!=null?(q.div_cp_div_total<0.25?'pos':'neg'):'neu'},
      {lbl:'Margem Líquida',val:q.margem_liquida!=null?fmtPct(q.margem_liquida):'—',ctx:'Lucratividade',cls:q.margem_liquida!=null?(q.margem_liquida>0.05?'pos':q.margem_liquida>0?'neu':'neg'):'neu'},
      {lbl:'ROIC',val:q.roic!=null?fmtPct(q.roic):'—',ctx:'Retorno capital',cls:q.roic!=null?(q.roic>0.08?'pos':q.roic>0?'neu':'neg'):'neu'},
      {lbl:'Caixa/Dív.CP',val:fmtN(q.caixa_div_cp),ctx:'Cobertura CP',cls:q.caixa_div_cp!=null?(q.caixa_div_cp>=1.5?'pos':q.caixa_div_cp>=1?'neu':'neg'):'neu'}
    ];
    const qualDims = [
      {lbl:'Vantagens Comp.', s:ql.vantagens_score||0, txt: ql.cmt_vantagens},
      {lbl:'Oportunidades',   s:ql.oportunidades_score||0, txt: ql.cmt_oportunidades},
      {lbl:'Vulnerabilidades',s:ql.vulnerabilidades_score||0, txt: ql.cmt_vulnerabilidades},
      {lbl:'Governança Corp.',s:ql.governanca_score||0, txt: ql.cmt_governanca},
      {lbl:'Setor',           s:ql.setor_score||0, txt: null},
      {lbl:'Rating',          s:ql.rating_score||0, txt: null}
    ];
    
    renderOmniModal(d, ctx, kpis, qualDims);
  }
}

function closeOmniModal(e) {
  if (e && e.target.id !== 'omni-overlay' && e.target.className !== 'omni-close') return;
  document.getElementById('omni-overlay').classList.remove('open');
  document.body.style.overflow = '';
}


// ══════════════════════════════════════════════════
// COMPARADOR — COMPARTILHADO
// ══════════════════════════════════════════════════
let selState = {
  privado: {a:null, b:null, chart:null},
  bancos:  {a:null, b:null, chart:null}
};

function setupAutocomplete(inputId, dropdownId, side, ctx) {
  const DATA = ctx === 'privado' ? DATA_PRIVADO : DATA_BANCOS;
  const inp = document.getElementById(inputId), dd = document.getElementById(dropdownId);
  inp.addEventListener('input', () => {
    const q = inp.value.toLowerCase().trim();
    if (!q) { dd.classList.remove('open'); dd.innerHTML=''; return; }
    const matches = DATA.filter(d => d.empresa.toLowerCase().includes(q)||(d.codigo||'').toLowerCase().includes(q)).slice(0,8);
    if (!matches.length) { dd.classList.remove('open'); return; }
    dd.innerHTML = matches.map(d=>`<div class="comp-opt" data-name="${d.empresa}"><span>${d.empresa}<br><span style="font-family:var(--mono);font-size:.58rem;color:var(--text3)">${d.codigo}</span></span><span class="comp-opt-rating"><span class="badge ${ratingCls(d.rating)}">${d.rating}</span></span></div>`).join('');
    dd.classList.add('open');
    dd.querySelectorAll('.comp-opt').forEach(opt => {
      opt.addEventListener('click', () => {
        const d = DATA.find(x => x.empresa === opt.dataset.name);
        if (d) selectCompany(side, d, ctx);
        dd.classList.remove('open'); inp.value='';
      });
    });
  });
  inp.addEventListener('blur', () => setTimeout(() => dd.classList.remove('open'), 150));
}

function selectCompany(side, d, ctx) {
  selState[ctx][side] = d;
  const sfx = ctx === 'bancos' ? 'b' : '';
  const chosenEl  = document.getElementById(sfx+'chosen-'+side);
  document.getElementById(sfx+'chosen-name-'+side).textContent = d.empresa;
  document.getElementById(sfx+'chosen-meta-'+side).textContent = `${d.codigo} · Ranking #${d.ranking} · ${d.rating}`;
  chosenEl.classList.add('visible');
  tryRenderComparador(ctx);
}

function clearSel(side, ctx) {
  selState[ctx][side] = null;
  const sfx = ctx === 'bancos' ? 'b' : '';
  document.getElementById(sfx+'chosen-'+side).classList.remove('visible');
  document.getElementById((ctx==='bancos'?'binp-':'inp-')+side).value = '';
  tryRenderComparador(ctx);
}

function sendToCompare(idx, ctx) {
  const DATA = ctx === 'privado' ? DATA_PRIVADO : DATA_BANCOS;
  const d = DATA[idx]; if (!d) return;
  const s = selState[ctx];
  if (!s.a) selectCompany('a', d, ctx);
  else if (!s.b && s.a.empresa !== d.empresa) selectCompany('b', d, ctx);
  else if (s.a && s.a.empresa === d.empresa) { if (!s.b || s.b.empresa !== d.empresa) selectCompany('b', d, ctx); }
  else { selectCompany('b', s.a, ctx); selectCompany('a', d, ctx); }
  closeModal();
  const compId = ctx === 'bancos' ? 'bcomparador' : 'comparador';
  setTimeout(() => document.getElementById(compId).scrollIntoView({behavior:'smooth',block:'start'}), 120);
}

function tryRenderComparador(ctx) {
  const sfx = ctx === 'bancos' ? 'b' : '';
  const body = document.getElementById(sfx+'comp-body');
  const ph   = document.getElementById(sfx+'comp-placeholder');
  const s    = selState[ctx];
  if (s.a && s.b) {
    ph.style.display = 'none';
    body.style.display = 'block';
    requestAnimationFrame(()=>requestAnimationFrame(()=>{ body.style.opacity='1'; }));
    renderComparador(ctx);
  } else {
    body.style.opacity='0';
    setTimeout(()=>{ body.style.display='none'; }, 300);
    ph.style.display='block';
    if (selState[ctx].chart) { selState[ctx].chart.destroy(); selState[ctx].chart=null; }
  }
}

function getQualDataPrivado(d) {
  const ql = d.qual||{};
  return [ql.vantagens_score||0, ql.oportunidades_score||0, ql.vulnerabilidades_score||0, ql.governanca_score||0, ql.setor_score||0, ql.rating_score||0];
}
function getQualDataBanco(d) {
  const ql = d.qual||{};
  return [ql.vantagens_score||0, ql.oportunidades_score||0, ql.vulnerabilidades_score||0, ql.governanca_score||0, ql.rating_score||0];
}

function renderComparador(ctx) {
  const sfx = ctx === 'bancos' ? 'b' : '';
  const a = selState[ctx].a, b = selState[ctx].b;
  document.getElementById(sfx+'legend-a').textContent = a.empresa;
  document.getElementById(sfx+'legend-b').textContent = b.empresa;

  const radarLabels = ctx === 'bancos'
    ? ['Vantagens','Oportunidades','Vulnerabilidades','Governança','Rating Agência']
    : ['Vantagens','Oportunidades','Vulnerabilidades','Governança','Setor','Rating'];
  const aData = ctx === 'bancos' ? getQualDataBanco(a) : getQualDataPrivado(a);
  const bData = ctx === 'bancos' ? getQualDataBanco(b) : getQualDataPrivado(b);

  const canvasId = ctx === 'bancos' ? 'bradar-chart' : 'radar-chart';
  if (selState[ctx].chart) selState[ctx].chart.destroy();
  const rctx = document.getElementById(canvasId).getContext('2d');

  const externalTooltipHandler = (context) => {
    let el = document.getElementById(sfx+'radar-tt');
    if (!el) {
      el = document.createElement('div'); el.id = sfx+'radar-tt';
      el.style.cssText = 'position:absolute;pointer-events:none;opacity:0;transition:opacity .18s;z-index:30;background:rgba(255,255,255,.72);backdrop-filter:blur(10px);border:1px solid rgba(31,40,57,.08);border-radius:10px;padding:10px 12px;box-shadow:0 8px 32px rgba(31,40,57,.12);font-family:Montserrat,sans-serif;font-size:.7rem;color:var(--navy);min-width:150px;transform:translate(-50%,-100%) translateY(-10px)';
      document.querySelector(`#${ctx==='bancos'?'b':''}comparador .comp-radar-wrap`).appendChild(el);
    }
    const tt = context.tooltip;
    if (tt.opacity === 0) { el.style.opacity='0'; return; }
    const idx = (tt.dataPoints||[])[0]?.dataIndex;
    const label = radarLabels[idx]||'';
    const va = aData[idx]??0, vb = bData[idx]??0;
    const diff = va-vb;
    const diffTxt = diff===0?'empate':(diff>0?`${a.empresa} +${diff}`:`${b.empresa} +${Math.abs(diff)}`);
    const diffColor = diff===0?'#718096':(diff>0?TEAL:GOLD);
    el.innerHTML = `<div style="font-size:.55rem;letter-spacing:.15em;text-transform:uppercase;color:var(--text3);margin-bottom:6px;font-weight:700">${label}</div><div style="display:flex;justify-content:space-between;gap:12px;margin-bottom:3px"><span style="color:${TEAL};font-weight:600">${a.empresa}</span><span style="font-family:'DM Mono',monospace;font-weight:700">${va}/5</span></div><div style="display:flex;justify-content:space-between;gap:12px;margin-bottom:8px"><span style="color:${GOLD};font-weight:600">${b.empresa}</span><span style="font-family:'DM Mono',monospace;font-weight:700">${vb}/5</span></div><div style="border-top:1px solid rgba(31,40,57,.08);padding-top:6px;font-size:.58rem;letter-spacing:.1em;text-transform:uppercase;color:${diffColor};font-weight:700;text-align:center">Δ ${diffTxt}</div>`;
    const wrap = document.querySelector(`#${ctx==='bancos'?'b':''}comparador .comp-radar-canvas-wrap`).getBoundingClientRect();
    const canvasEl = context.chart.canvas.getBoundingClientRect();
    el.style.left = (tt.caretX + canvasEl.left - wrap.left)+'px';
    el.style.top  = (tt.caretY + canvasEl.top  - wrap.top )+'px';
    el.style.opacity = '1';
  };

  selState[ctx].chart = new Chart(rctx, {
    type:'radar',
    data:{ labels:radarLabels, datasets:[
      {label:a.empresa,data:aData,borderColor:TEAL,backgroundColor:'rgba(0,103,123,0.12)',borderWidth:2,pointBackgroundColor:TEAL,pointBorderColor:'#fff',pointRadius:4,pointHoverRadius:6},
      {label:b.empresa,data:bData,borderColor:GOLD,backgroundColor:'rgba(182,157,116,0.12)',borderWidth:2,pointBackgroundColor:GOLD,pointBorderColor:'#fff',pointRadius:4,pointHoverRadius:6}
    ]},
    options:{responsive:true,maintainAspectRatio:false,animation:{duration:700,easing:'easeInOutQuart'},interaction:{mode:'index',intersect:false},scales:{r:{min:0,max:5,ticks:{stepSize:1,display:false},grid:{color:'rgba(31,40,57,0.08)'},angleLines:{color:'rgba(31,40,57,0.08)'},pointLabels:{font:{family:"'Montserrat',sans-serif",size:10,weight:'600'},color:'#4a5568'}}},plugins:{legend:{display:false},tooltip:{enabled:false,external:externalTooltipHandler}}}
  });

  // Scores panel
  const scoreDefs = [{lbl:'Score Quant.',keyA:'scoreQuantitativo',higher:true},{lbl:'Score Qual.',keyA:'scoreQualitativo',higher:true},{lbl:'Score Total',keyA:'scoreTotal',higher:true}];
  document.getElementById(sfx+'comp-scores-panel').innerHTML = scoreDefs.map(s => {
    const va=a[s.keyA],vb=b[s.keyA];
    const winA=va>vb,winB=vb>va;
    const maxV=Math.max(va,vb)||1;
    const hA=Math.round((va/maxV)*50),hB=Math.round((vb/maxV)*50);
    return `<div class="comp-score-row"><div class="comp-score-col"><div class="comp-score-lbl-tiny" style="color:${TEAL}">${a.empresa}</div><div class="comp-score-val${winA?' winner':''}" style="color:${winA?TEAL:'var(--text2)'}">${va.toFixed(1)}</div></div><div class="comp-score-bar-center"><div class="comp-score-bar-lbl">${s.lbl}</div><div class="comp-duel-bar"><div class="comp-duel-fill-a" style="height:${hA}px"></div><div class="comp-duel-fill-b" style="height:${hB}px"></div></div></div><div class="comp-score-col right"><div class="comp-score-lbl-tiny" style="color:${GOLD}">${b.empresa}</div><div class="comp-score-val${winB?' winner':''}" style="color:${winB?GOLD:'var(--text2)'}">${vb.toFixed(1)}</div></div></div>`;
  }).join('');

  // Cabo de guerra KPIs
  const qa = a.quant || {};
  const qb = b.quant || {};
  let tugKpis = [];

  if (ctx === 'bancos') {
    tugKpis = [
      {name:'Basileia',ctx:'maior = melhor',key:'basileia',fmt:v=>fmtPctDirect(v),higher:true},
      {name:'Alavancagem',ctx:'maior = melhor',key:'alavancagem',fmt:v=>fmtPctDirect(v),higher:true},
      {name:'Imobilização',ctx:'menor = melhor',key:'imobilizacao',fmt:v=>fmtPctDirect(v),higher:false},
      {name:'Provisão/Cart.',ctx:'menor = melhor',key:'provisao',fmt:v=>fmtPctDirect(v),higher:false},
      {name:'Eficiência',ctx:'menor = melhor',key:'eficiencia',fmt:v=>fmtPctDirect(v),higher:false},
      {name:'Margem (NIM)',ctx:'maior = melhor',key:'margem',fmt:v=>fmtPctDirect(v),higher:true},
      {name:'ROE',ctx:'maior = melhor',key:'roe',fmt:v=>fmtPctDirect(v),higher:true},
    ];
  } else {
    tugKpis = [
      {name:'Dív.Líq./EBITDA',ctx:'menor = melhor',key:'div_liq_ebitda',fmt:v=>fmtN(v),higher:false},
      {name:'EBITDA/Desp.Fin.',ctx:'maior = melhor',key:'ebitda_desp_fin',fmt:v=>fmtN(v),higher:true},
      {name:'Liquidez Corrente',ctx:'maior = melhor',key:'liquidez_corrente',fmt:v=>fmtN(v),higher:true},
      {name:'Margem Líquida',ctx:'maior = melhor',key:'margem_liquida',fmt:v=>fmtPct(v),higher:true},
      {name:'ROIC',ctx:'maior = melhor',key:'roic',fmt:v=>fmtPct(v),higher:true},
      {name:'Caixa/Dív.CP',ctx:'maior = melhor',key:'caixa_div_cp',fmt:v=>fmtN(v),higher:true},
      {name:'Estrutura Capital',ctx:'menor = melhor',key:'estrutura_capital',fmt:v=>fmtPct(v),higher:false},
      {name:'Dív.CP/Dív.Total',ctx:'menor = melhor',key:'div_cp_div_total',fmt:v=>fmtPct(v),higher:false},
    ];
  }

  document.getElementById(sfx+'comp-kpi-table').innerHTML = tugKpis.map(k => {
    const va = qa[k.key];
    const vb = qb[k.key];
    const bothValid = va != null && vb != null;
    const winA = bothValid && (k.higher ? va > vb : va < vb);
    const winB = bothValid && (k.higher ? vb > va : vb < va);
    
    return `<div class="comp-kpi-row">
      <div class="comp-kpi-val-a${winA?' win':''}">${k.fmt(va)}</div>
      <div class="comp-kpi-center">
        <div class="comp-kpi-name">${k.name}</div>
        <div class="comp-kpi-ctx">${k.ctx}</div>
      </div>
      <div class="comp-kpi-val-b${winB?' win':''}">${k.fmt(vb)}</div>
    </div>`;
  }).join('');
}

setupAutocomplete('inp-a',  'dd-a',  'a', 'privado');
setupAutocomplete('inp-b',  'dd-b',  'b', 'privado');
setupAutocomplete('binp-a', 'bdd-a', 'a', 'bancos');
setupAutocomplete('binp-b', 'bdd-b', 'b', 'bancos');

// ══════════════════════════════════════════════════
// ANÁLISE SETORIAL — PRIVADO
// ══════════════════════════════════════════════════
let secChart = null;
const sectorKpis = [
  {id:'margem_liquida',label:'Margem Líquida',isPct:true,asc:false},
  {id:'roic',label:'ROIC',isPct:true,asc:false},
  {id:'div_liq_ebitda',label:'Dív.Líq/EBITDA',isPct:false,asc:true},
  {id:'liquidez_corrente',label:'Liquidez Corrente',isPct:false,asc:false}
];
let currentSectorKpi = 'margem_liquida';
let currentSectorData = [];

function initSectorAnalysis() {
  const secs = [...new Set(DATA_PRIVADO.map(d=>d.qual&&d.qual.setor).filter(Boolean))].sort();
  const sel = document.getElementById('sector-select');
  secs.forEach(s => { const opt=document.createElement('option'); opt.value=s; opt.textContent=s; sel.appendChild(opt); });
  sel.addEventListener('change', e => {
    const s = e.target.value;
    const b=document.getElementById('sec-body'), p=document.getElementById('sec-placeholder');
    if (!s) { b.style.opacity='0'; setTimeout(()=>{b.style.display='none';p.style.display='block';},300); return; }
    p.style.display='none'; b.style.display='block';
    requestAnimationFrame(()=>requestAnimationFrame(()=>b.style.opacity='1'));
    currentSectorData = DATA_PRIVADO.filter(d=>d.qual&&d.qual.setor===s&&d.quant!=null);
    let tableData=[...currentSectorData].sort((a,b)=>b.scoreTotal-a.scoreTotal);
    const avg=tableData.reduce((a,c)=>a+c.scoreTotal,0)/(tableData.length||1);
    const best=tableData[0];
    document.getElementById('sec-stats').innerHTML=`<div class="stat-box s-total" style="flex:1;padding:12px 16px"><div class="stat-num" style="font-size:1.5rem">${tableData.length}</div><div class="stat-lbl">Empresas Analisadas</div></div><div class="stat-box s-anal" style="flex:1;padding:12px 16px"><div class="stat-num c-gold" style="font-size:1.5rem">${avg.toFixed(1)}</div><div class="stat-lbl">Score Médio</div></div><div class="stat-box s-aprov" style="flex:1;padding:12px 16px"><div class="stat-num c-teal" style="font-size:1.5rem">${best?best.scoreTotal.toFixed(1):'—'}</div><div class="stat-lbl">Líder (${best?best.empresa:'—'})</div></div>`;
    document.getElementById('sec-tbl-body').innerHTML=tableData.map((d,i)=>{const idx=DATA_PRIVADO.findIndex(x=>x.empresa===d.empresa);return`<tr onclick="openModal(${idx},'privado')"><td><div class="rank-cell">${medal(i+1)}</div></td><td><div class="co-name">${d.empresa}</div><div class="co-code">${d.codigo}</div></td><td>${statusHTML(d.status)}</td><td><span class="bar-n" style="font-size:.75rem">${d.scoreQuantitativo.toFixed(1)}</span></td><td><span class="bar-n" style="font-size:.75rem">${d.scoreQualitativo.toFixed(1)}</span></td><td><span class="bar-n" style="font-weight:700;color:var(--teal);font-size:.8rem">${d.scoreTotal.toFixed(1)}</span></td><td><span class="badge ${ratingCls(d.rating)}">${d.rating}</span></td></tr>`;}).join('');
    renderSectorChart();
  });
}

function renderSectorChart() {
  const kpiDef = sectorKpis.find(k => k.id === currentSectorKpi);
  let chartData = [...currentSectorData].filter(d => d.quant[kpiDef.id] != null);
  chartData.sort((a,b) => kpiDef.asc ? a.quant[kpiDef.id] - b.quant[kpiDef.id] : b.quant[kpiDef.id] - a.quant[kpiDef.id]);
  
  document.getElementById('sec-kpi-tabs').innerHTML = sectorKpis.map(k => `<button class="kpi-tab ${k.id === currentSectorKpi ? 'active' : ''}" onclick="changeSectorKpi('${k.id}')">${k.label}</button>`).join('');
  
  const labels = chartData.map(d => d.empresa.length > 12 ? d.empresa.substring(0,10) + '...' : d.empresa);
  const vals = chartData.map(d => d.quant[kpiDef.id]);
  const colors = chartData.map(d => {
    if (d.status === 'Aprovado') return '#00677b';
    if (d.status === 'Em análise') return '#b69d74';
    return '#922b21';
  });

  const datasets = [{
    type: 'bar',
    label: kpiDef.label,
    data: vals,
    backgroundColor: colors,
    borderRadius: 4,
    order: 1 // Barras ficam por baixo
  }];

  // INJEÇÃO DA LINHA DE BENCHMARK
  if (privBenchMode !== 'none' && vals.length > 0) {
    const valLine = privBenchMode === 'media' ? calcMean(vals) : calcMedian(vals);
    datasets.push({
      type: 'line',
      label: privBenchMode === 'media' ? 'Média do Setor' : 'Mediana do Setor',
      data: Array(vals.length).fill(valLine),
      borderColor: '#1f2839', // Navy line
      borderWidth: 2,
      borderDash: [5, 5],
      pointRadius: 0,
      pointHoverRadius: 0,
      fill: false,
      order: 0 // Linha fica por cima das barras
    });
  }

  if (secChart) secChart.destroy();
  const ctx2 = document.getElementById('sector-chart').getContext('2d');
  secChart = new Chart(ctx2, {
    data: { labels, datasets },
    options: {
      responsive: true,
      maintainAspectRatio: false,
      animation: { duration: 600, easing: 'easeOutQuart' },
      scales: {
        y: { grid: { color: 'rgba(31,40,57,.05)' }, ticks: { callback: val => kpiDef.isPct ? (val * 100).toFixed(0) + '%' : val.toFixed(1) + 'x' } },
        x: { grid: { display: false }, ticks: { font: { family: "'Montserrat',sans-serif", size: 10, weight: 600 } } }
      },
      plugins: {
        legend: { display: false },
        tooltip: {
          callbacks: {
            label: context => {
              const labelPrefix = context.dataset.type === 'line' ? context.dataset.label : kpiDef.label;
              return `${labelPrefix}: ` + (kpiDef.isPct ? fmtPct(context.raw) : fmtN(context.raw) + 'x');
            }
          }
        }
      }
    }
  });
}
window.changeSectorKpi = id => { currentSectorKpi=id; renderSectorChart(); };
initSectorAnalysis();

// ══════════════════════════════════════════════════
// ANÁLISE SETORIAL — BANCOS
// ══════════════════════════════════════════════════
let bsecChart=null;
const bankKpis=[
  {id:'basileia',  label:'Basileia',    isPct:true,asc:false},
  {id:'roe',       label:'ROE',         isPct:true,asc:false},
  {id:'eficiencia',label:'Eficiência',  isPct:true,asc:true},
  {id:'margem',    label:'Margem (NIM)',isPct:true,asc:false},
  {id:'provisao',  label:'Provisão',    isPct:true,asc:true},
];
let currentBankKpi='basileia';
let currentBankSectorData=[];

function initBankSectorAnalysis(){
  const secs=[...new Set(DATA_BANCOS.map(d=>d.setor).filter(Boolean))].sort();
  const sel=document.getElementById('bsector-select');
  secs.forEach(s=>{const opt=document.createElement('option');opt.value=s;opt.textContent=s;sel.appendChild(opt);});
  sel.addEventListener('change',e=>{
    const s=e.target.value;
    const b=document.getElementById('bsec-body'),p=document.getElementById('bsec-placeholder');
    if(!s){b.style.opacity='0';setTimeout(()=>{b.style.display='none';p.style.display='block';},300);return;}
    p.style.display='none';b.style.display='block';
    requestAnimationFrame(()=>requestAnimationFrame(()=>b.style.opacity='1'));
    currentBankSectorData=DATA_BANCOS.filter(d=>d.setor===s);
    let tableData=[...currentBankSectorData].sort((a,b)=>b.scoreTotal-a.scoreTotal);
    const avg=tableData.reduce((a,c)=>a+c.scoreTotal,0)/(tableData.length||1);
    const best=tableData[0];
    document.getElementById('bsec-stats').innerHTML=`<div class="stat-box s-total" style="flex:1;padding:12px 16px"><div class="stat-num" style="font-size:1.5rem">${tableData.length}</div><div class="stat-lbl">Bancos Analisados</div></div><div class="stat-box s-anal" style="flex:1;padding:12px 16px"><div class="stat-num c-gold" style="font-size:1.5rem">${avg.toFixed(1)}</div><div class="stat-lbl">Score Médio</div></div><div class="stat-box s-aprov" style="flex:1;padding:12px 16px"><div class="stat-num c-teal" style="font-size:1.5rem">${best?best.scoreTotal.toFixed(1):'—'}</div><div class="stat-lbl">Líder (${best?best.empresa:'—'})</div></div>`;
    document.getElementById('bsec-tbl-body').innerHTML=tableData.map((d,i)=>{const idx=DATA_BANCOS.findIndex(x=>x.empresa===d.empresa);return`<tr onclick="openModal(${idx},'bancos')"><td><div class="rank-cell">${medal(i+1)}</div></td><td><div class="co-name">${d.empresa}</div><div class="co-code">${d.codigo}</div></td><td>${statusHTML(d.status)}</td><td><span class="bar-n" style="font-size:.75rem">${d.scoreQuantitativo.toFixed(1)}</span></td><td><span class="bar-n" style="font-size:.75rem">${d.scoreQualitativo.toFixed(1)}</span></td><td><span class="bar-n" style="font-weight:700;color:var(--teal);font-size:.8rem">${d.scoreTotal.toFixed(1)}</span></td><td><span class="badge ${ratingCls(d.rating)}">${d.rating}</span></td></tr>`;}).join('');
    renderBankSectorChart();
  });
}

function renderBankSectorChart(){
  const kpiDef = bankKpis.find(k => k.id === currentBankKpi);
  let chartData = [...currentBankSectorData].filter(d => d.quant[kpiDef.id] != null);
  chartData.sort((a,b) => kpiDef.asc ? a.quant[kpiDef.id] - b.quant[kpiDef.id] : b.quant[kpiDef.id] - a.quant[kpiDef.id]);
  
  document.getElementById('bsec-kpi-tabs').innerHTML = bankKpis.map(k => `<button class="kpi-tab ${k.id === currentBankKpi ? 'active' : ''}" onclick="changeBankKpi('${k.id}')">${k.label}</button>`).join('');
  
  const labels = chartData.map(d => d.empresa.length > 12 ? d.empresa.substring(0,10) + '...' : d.empresa);
  const vals = chartData.map(d => d.quant[kpiDef.id]);
  const colors = chartData.map(d => {
    if (d.status === 'Aprovado') return '#00677b';
    if (d.status === 'Em análise') return '#b69d74';
    if (d.status === 'FGC') return '#7a6518'; 
    return '#922b21';
  });

  const datasets = [{
    type: 'bar',
    label: kpiDef.label,
    data: vals,
    backgroundColor: colors,
    borderRadius: 4,
    order: 1
  }];

  // INJEÇÃO DA LINHA DE BENCHMARK (BANCOS)
  if (bankBenchMode !== 'none' && vals.length > 0) {
    const valLine = bankBenchMode === 'media' ? calcMean(vals) : calcMedian(vals);
    datasets.push({
      type: 'line',
      label: bankBenchMode === 'media' ? 'Média do Segmento' : 'Mediana do Segmento',
      data: Array(vals.length).fill(valLine),
      borderColor: '#1f2839',
      borderWidth: 2,
      borderDash: [5, 5],
      pointRadius: 0,
      pointHoverRadius: 0,
      fill: false,
      order: 0
    });
  }

  if (bsecChart) bsecChart.destroy();
  const ctx3 = document.getElementById('bsector-chart').getContext('2d');
  bsecChart = new Chart(ctx3, {
    data: { labels, datasets },
    options: {
      responsive: true,
      maintainAspectRatio: false,
      animation: { duration: 600, easing: 'easeOutQuart' },
      scales: {
        y: { grid: { color: 'rgba(31,40,57,.05)' }, ticks: { callback: val => kpiDef.isPct ? (val * 100).toFixed(1) + '%' : val.toFixed(2) + 'x' } },
        x: { grid: { display: false }, ticks: { font: { family: "'Montserrat',sans-serif", size: 10, weight: 600 } } }
      },
      plugins: {
        legend: { display: false },
        tooltip: {
          callbacks: {
            label: context => {
              const labelPrefix = context.dataset.type === 'line' ? context.dataset.label : kpiDef.label;
              return `${labelPrefix}: ` + (kpiDef.isPct ? fmtPctDirect(context.raw) : fmtN(context.raw) + 'x');
            }
          }
        }
      }
    }
  });
}
window.changeBankKpi = id => { currentBankKpi=id; renderBankSectorChart(); };
initBankSectorAnalysis();

// ══════════════════════════════════════════════════
// INITIAL RENDER
// ══════════════════════════════════════════════════
renderTable('privado');
renderRatingsCard();
renderTable('bancos');

// ══════════════════════════════════════════════════
// IntersectionObserver
// ══════════════════════════════════════════════════
const io=new IntersectionObserver(entries=>{entries.forEach(e=>{if(e.isIntersecting){e.target.classList.add('visible');io.unobserve(e.target)}});},{threshold:.07});
document.querySelectorAll('.reveal').forEach(el=>io.observe(el));
document.addEventListener('scroll',()=>{const pct=scrollY/(document.body.scrollHeight-innerHeight)*100;document.getElementById('progress').style.width=pct+'%'});

// ══════════════════════════════════════════════════
// Custom cursor
// ══════════════════════════════════════════════════
const cur=document.getElementById('cur'),ring=document.getElementById('cur-ring');
let mx=0,my=0,rx=0,ry=0;
document.addEventListener('mousemove',e=>{mx=e.clientX;my=e.clientY;});
(function anim(){rx+=(mx-rx)*.12;ry+=(my-ry)*.12;cur.style.left=mx+'px';cur.style.top=my+'px';ring.style.left=rx+'px';ring.style.top=ry+'px';requestAnimationFrame(anim)})();
document.querySelectorAll('button,th,tbody tr,.pod,.comp-opt,.comp-input, .omni-qdim, .omni-close, .modal-close').forEach(el=>{
  el.addEventListener('mouseenter',()=>{cur.style.width='5px';cur.style.height='5px';ring.style.width='40px';ring.style.height='40px';ring.style.borderColor='var(--teal)';ring.style.opacity='.7'});
  el.addEventListener('mouseleave',()=>{cur.style.width='7px';cur.style.height='7px';ring.style.width='28px';ring.style.height='28px';ring.style.borderColor='var(--teal)';ring.style.opacity='.4'});
});

// ══════════════════════════════════════════════════
// Particle canvas
// ══════════════════════════════════════════════════
(function(){
  const cv=document.getElementById('canvas'),ctx=cv.getContext('2d');
  let W,H,pts;
  const init=()=>{W=cv.width=innerWidth;H=cv.height=innerHeight;pts=Array.from({length:55},()=>({x:Math.random()*W,y:Math.random()*H,r:Math.random()*.9+.2,vx:(Math.random()-.5)*.14,vy:(Math.random()-.5)*.14,o:Math.random()*.2+.04,c:Math.random()>.5?'31,40,57':'0,103,123'}))};
  addEventListener('resize',init);init();
  (function draw(){ctx.clearRect(0,0,W,H);pts.forEach(p=>{p.x=(p.x+p.vx+W)%W;p.y=(p.y+p.vy+H)%H;ctx.beginPath();ctx.arc(p.x,p.y,p.r,0,Math.PI*2);ctx.fillStyle=`rgba(${p.c},${p.o})`;ctx.fill()});requestAnimationFrame(draw)})();
})();
</script>
</body>
</html>"""


# ─────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────

def main():
    print("\n══════════════════════════════════════════════════")
    print("  Douro Capital — Gerador de Scorecard de Crédito")
    print("══════════════════════════════════════════════════\n")

    # ── Crédito Privado ──────────────────────────────────────
    dataset_privado = []
    if not os.path.exists(EXCEL_PATH_PRIVADO):
        print(f"[AVISO] Planilha Crédito Privado não encontrada:\n  {EXCEL_PATH_PRIVADO}")
        print("  O scorecard será gerado apenas com dados bancários.\n")
    else:
        try:
            dataset_privado = build_dataset(EXCEL_PATH_PRIVADO)
        except Exception as e:
            print(f"[ERRO] Falha ao ler planilha Crédito Privado: {e}")
            sys.exit(1)

    # ── Crédito Bancário ─────────────────────────────────────
    dataset_bancos = []
    if not os.path.exists(EXCEL_PATH_BANCOS):
        print(f"[AVISO] Planilha Crédito Bancário não encontrada:\n  {EXCEL_PATH_BANCOS}")
        print("  O scorecard será gerado apenas com dados privados.\n")
    else:
        try:
            
            dataset_bancos = build_bank_dataset(EXCEL_PATH_BANCOS)
        except Exception as e:
            print(f"[ERRO] Falha ao ler planilha Crédito Bancário: {e}")
            sys.exit(1)

    if not dataset_privado and not dataset_bancos:
        print("[ERRO] Nenhuma planilha encontrada. Verifique os caminhos no início do script.")
        sys.exit(1)

    now = datetime.datetime.now()
    MESES_PT = {
        1:"Janeiro",2:"Fevereiro",3:"Março",4:"Abril",
        5:"Maio",6:"Junho",7:"Julho",8:"Agosto",
        9:"Setembro",10:"Outubro",11:"Novembro",12:"Dezembro",
    }
    month_label  = f"{MESES_PT[now.month]} de {now.year}"
    generated_at = now.strftime("%d/%m/%Y %H:%M")

    # Stats — privado
    total_p      = len(dataset_privado)
    aprovadas_p  = sum(1 for d in dataset_privado if d.get("status") == "Aprovado")
    em_analise_p = sum(1 for d in dataset_privado if d.get("status") == "Em análise")
    reprovadas_p = sum(1 for d in dataset_privado if d.get("status") == "Reprovado")

    # Stats — bancos
    total_b      = len(dataset_bancos)
    aprovados_b  = sum(1 for d in dataset_bancos if d.get("status") == "Aprovado")
    em_analise_b = sum(1 for d in dataset_bancos if d.get("status") == "Em análise")
    reprovados_b = sum(1 for d in dataset_bancos if d.get("status") in ("Reprovado", "FGC"))

    tmpl   = Template(HTML_TEMPLATE)
    logo_src = image_to_data_uri(LOGO_PATH)
    if not logo_src:
        print(f"[AVISO] Logo não encontrada ou inválida: {LOGO_PATH}")
        print("  O HTML será gerado sem logo embutida.\n")
    output = tmpl.render(
        data_privado_json = json.dumps(dataset_privado, ensure_ascii=False),
        data_bancos_json  = json.dumps(dataset_bancos,  ensure_ascii=False),
        total_p      = total_p,
        aprovadas_p  = aprovadas_p,
        em_analise_p = em_analise_p,
        reprovadas_p = reprovadas_p,
        total_b      = total_b,
        aprovados_b  = aprovados_b,
        em_analise_b = em_analise_b,
        reprovados_b = reprovados_b,
        month_label  = month_label,
        generated_at = generated_at,
        logo_src     = logo_src,
    )

    with open(OUTPUT_HTML, "w", encoding="utf-8") as f:
        f.write(output)

    size_kb = os.path.getsize(OUTPUT_HTML) / 1024
    print(f"\n  ✔ HTML gerado: {OUTPUT_HTML}")
    print(f"  ✔ Tamanho: {size_kb:.1f} KB")
    print(f"  ✔ Crédito Privado: {total_p} empresas  |  Aprovadas: {aprovadas_p}  |  Em análise: {em_analise_p}  |  Reprovadas: {reprovadas_p}")
    print(f"  ✔ Crédito Bancário: {total_b} bancos    |  Aprovados: {aprovados_b}  |  Em análise: {em_analise_b}  |  Reprovados/FGC: {reprovados_b}")
    print(f"\n  Abra '{os.path.basename(OUTPUT_HTML)}' no navegador.\n")


if __name__ == "__main__":
    main()